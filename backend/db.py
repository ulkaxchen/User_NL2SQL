from __future__ import annotations

import csv
import difflib
import json
import logging
import os
import re
from collections import deque
from io import BytesIO, StringIO
from datetime import datetime
from pathlib import Path
from typing import Any

# 在读取任何环境变量之前加载项目根目录的 .env（与 uvicorn / python -m backend 工作目录无关）
_REPO_ROOT = Path(__file__).resolve().parent.parent
try:
    from dotenv import load_dotenv

    load_dotenv(_REPO_ROOT / ".env")
    load_dotenv(_REPO_ROOT / ".env.local")  # 本地覆盖，可不提交 git
except ImportError:
    pass

import pymysql
from pymysql.cursors import DictCursor

# 执行层默认最大行数（防 OOM）；导出接口传入 max_rows=-1 不截断
EXECUTE_SQL_MAX_ROWS = min(max(int(os.environ.get("USER_RAG_EXECUTE_MAX_ROWS", "1000")), 1), 500_000)
# 探查类工具：大表上「每列一次全表扫描」成本极高，限制每轮最多处理的列数（可调大，查询会变慢）
USER_RAG_KEYWORD_MAX_COLUMNS = min(max(int(os.environ.get("USER_RAG_KEYWORD_MAX_COLUMNS", "20")), 4), 80)
USER_RAG_PROFILE_TABLE_MAX_COLUMNS = min(max(int(os.environ.get("USER_RAG_PROFILE_TABLE_MAX_COLUMNS", "24")), 4), 80)

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .country_match import country_literal_candidates, resolve_country_intent

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DEFAULT_XLSX_PATH = BASE_DIR.parent / "总用户数据下钻列表_20260318.xlsx"
USERS_TABLE = "users"
# 不参与业务查询的系统表（仅日志；业务表来自 information_schema）
INTERNAL_TABLES = frozenset({"query_logs"})
_INTERNAL_TABLE_NAMES_LOWER = {x.lower() for x in INTERNAL_TABLES}
MYSQL_RESERVED_SCHEMAS = frozenset({"information_schema", "mysql", "performance_schema", "sys"})

MYSQL_HOST = (os.environ.get("USER_RAG_MYSQL_HOST") or os.environ.get("MYSQL_HOST") or "127.0.0.1").strip()
MYSQL_PORT = int(os.environ.get("USER_RAG_MYSQL_PORT") or os.environ.get("MYSQL_PORT") or "3306")
MYSQL_USER = (os.environ.get("USER_RAG_MYSQL_USER") or os.environ.get("MYSQL_USER") or "root").strip()
MYSQL_PASSWORD = os.environ.get("USER_RAG_MYSQL_PASSWORD") or os.environ.get("MYSQL_PASSWORD") or ""
MYSQL_DATABASE = (os.environ.get("USER_RAG_MYSQL_DATABASE") or os.environ.get("MYSQL_DATABASE") or "").strip()


def get_mysql_database_label() -> str:
    """供 ImportResponse / 展示：当前连接的库名。"""
    return MYSQL_DATABASE or "mysql:(未配置 MYSQL_DATABASE)"


def validate_business_table_name(name: str) -> str:
    """导入目标表名：字母数字下划线，且不得与系统表冲突。"""
    t = str(name or "").strip()
    if not t:
        raise ValueError("表名不能为空")
    if not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", t):
        raise ValueError("表名须为字母、数字、下划线，且不能以数字开头")
    tl = t.lower()
    if tl in {x.lower() for x in INTERNAL_TABLES}:
        raise ValueError(f"不允许使用保留表名: {t}")
    if tl in MYSQL_RESERVED_SCHEMAS:
        raise ValueError(f"不允许使用保留表名: {t}")
    return t


FORBIDDEN_SQL = (
    "insert",
    "update",
    "delete",
    "drop",
    "alter",
    "create",
    "attach",
    "pragma",
    "load_file",
    "into outfile",
    "into dumpfile",
)


def ensure_directories() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def get_connection() -> pymysql.Connection:
    if not MYSQL_DATABASE:
        raise RuntimeError("未配置 MySQL 数据库名：请设置 USER_RAG_MYSQL_DATABASE 或 MYSQL_DATABASE")
    return pymysql.connect(
        host=MYSQL_HOST,
        port=MYSQL_PORT,
        user=MYSQL_USER,
        password=MYSQL_PASSWORD,
        database=MYSQL_DATABASE,
        charset="utf8mb4",
        cursorclass=DictCursor,
        autocommit=False,
    )


def _conn_exec(conn: pymysql.Connection, sql: str, params: tuple | list | None = None) -> pymysql.cursors.Cursor:
    """执行 SQL。无占位符时必须 **不要** 传空元组：PyMySQL 仍会对 query 做 %% 风格替换，
    动态 SQL 里的 LIKE '%词%' 会被当成 format，触发 not enough arguments for format string。"""
    cur = conn.cursor()
    if not params:
        cur.execute(sql)
    else:
        cur.execute(sql, params)
    return cur


def _conn_executemany(conn: pymysql.Connection, sql: str, seq: list) -> None:
    cur = conn.cursor()
    cur.executemany(sql, seq)


def _quote_identifier(identifier: str) -> str:
    escaped = str(identifier).replace("`", "``")
    return f"`{escaped}`"


def _dict_cursor_cell(row: Any, *preferred_keys: str) -> Any:
    """DictCursor 行按列名取值；禁止用 row[0]（会得到 KeyError(0)，上层只看到 error \"0\"）。"""
    if not row:
        return None
    if not isinstance(row, dict):
        return row[0]
    for k in preferred_keys:
        if k and k in row:
            return row[k]
    if len(row) == 1:
        return next(iter(row.values()))
    return None


def _row_to_ordered_values(row: Any, columns: list[str]) -> list[Any]:
    """DictCursor 下行是 dict：`list(row)` 只会得到列名，必须用列顺序取 value。"""
    if isinstance(row, dict):
        return [row.get(c) for c in columns]
    return list(row)


def _table_exists(table_name: str) -> bool:
    if not MYSQL_DATABASE:
        return False
    with get_connection() as conn:
        row = _conn_exec(
            conn,
            """
            SELECT 1 AS ok FROM information_schema.tables
            WHERE table_schema = %s AND table_name = %s AND table_type = 'BASE TABLE'
            LIMIT 1
            """,
            (MYSQL_DATABASE, table_name),
        ).fetchone()
    return row is not None


def _table_names_from_show_full_tables(rows: list[dict[str, Any]]) -> list[str]:
    """解析 SHOW FULL TABLES 结果（列名形如 Tables_in_<库名>）。"""
    if not rows:
        return []
    keys = list(rows[0].keys())
    name_key = next((k for k in keys if str(k).startswith("Tables_in_")), keys[0])
    out: list[str] = []
    for row in rows:
        ttype = str(row.get("Table_type") or row.get("table_type") or "BASE TABLE").upper()
        if ttype != "BASE TABLE":
            continue
        raw = row.get(name_key)
        if raw is not None and str(raw).strip():
            out.append(str(raw).strip())
    return sorted(set(out))


def queryable_table_names() -> list[str]:
    """当前库中允许探查与 SELECT 的表名（已排除日志表）。"""
    if not MYSQL_DATABASE:
        logger.warning("queryable_table_names: 未设置 MYSQL_DATABASE / USER_RAG_MYSQL_DATABASE，表列表为空")
        return []
    rows: list[dict[str, Any]] = []
    try:
        with get_connection() as conn:
            rows = _conn_exec(
                conn,
                """
                SELECT TABLE_NAME AS name FROM information_schema.tables
                WHERE table_schema = %s AND table_type = 'BASE TABLE'
                ORDER BY TABLE_NAME
                """,
                (MYSQL_DATABASE,),
            ).fetchall()
    except pymysql.Error as exc:
        logger.warning("queryable_table_names: information_schema 查询失败，尝试 SHOW FULL TABLES: %s", exc)
        try:
            with get_connection() as conn:
                rows = _conn_exec(conn, "SHOW FULL TABLES").fetchall()
                rows = [{"name": n} for n in _table_names_from_show_full_tables(rows)]
        except pymysql.Error as exc2:
            logger.exception("queryable_table_names: SHOW FULL TABLES 也失败: %s", exc2)
            return []
    out: list[str] = []
    for row in rows:
        name = str(row["name"])
        if name.lower() in _INTERNAL_TABLE_NAMES_LOWER:
            continue
        out.append(name)
    return out


def is_queryable_table(table_name: str) -> bool:
    t = str(table_name or "").strip()
    if not t or t.lower() in _INTERNAL_TABLE_NAMES_LOWER:
        return False
    return _table_exists(t)


def resolve_table_name(table: str | None) -> str:
    return str(table or "").strip() or USERS_TABLE


def column_names_for_table(table_name: str) -> list[str]:
    return [c["name"] for c in list_columns(table_name)]


_COUNTRY_PHONE_PREFIX_HINTS: dict[str, tuple[str, ...]] = {
    "indonesia": ("+62",),
    "india": ("+91",),
    "china": ("+86",),
    "japan": ("+81",),
    "south korea": ("+82",),
    "vietnam": ("+84",),
    "thailand": ("+66",),
    "malaysia": ("+60",),
    "singapore": ("+65",),
    "philippines": ("+63",),
    "united states": ("+1",),
    "saudi arabia": ("+966",),
    "united arab emirates": ("+971",),
}

_FILTER_NAME_HINTS: dict[str, dict[str, tuple[str, ...]]] = {
    "country": {
        "strong": ("国家", "地区", "区域", "城市", "省", "country", "nation", "region", "area", "market", "city", "province"),
        "fallback": ("联系电话", "电话", "手机", "mobile", "phone", "邮箱", "email", "地址", "address", "标签", "tag"),
        "negative": ("平台", "渠道", "platform", "channel", "版本", "version", "时间", "日期", "date", "time"),
    },
    "platform": {
        "strong": ("平台", "渠道", "platform", "channel", "device", "os", "terminal", "终端"),
        "fallback": ("app", "客户端", "来源", "source", "版本", "version"),
        "negative": ("国家", "地区", "country", "region"),
    },
    "category": {
        "strong": ("状态", "类型", "角色", "身份", "标签", "分类", "status", "type", "role", "identity", "tag", "category", "segment"),
        "fallback": ("客户", "customer", "group", "level", "grade", "class"),
        "negative": ("时间", "日期", "date", "time", "数量", "sum", "qty"),
    },
    "time": {
        "strong": ("时间", "日期", "年月", "月", "日", "time", "date", "created", "updated", "active", "register", "month", "day", "year"),
        "fallback": ("dt", "timestamp"),
        "negative": ("平台", "渠道", "platform", "channel"),
    },
    "id": {
        "strong": ("id", "uid", "编号", "编码", "单号", "order_no", "code"),
        "fallback": ("用户", "客户", "订单", "物料", "user", "customer", "order", "matnr", "sku", "item"),
        "negative": ("时间", "日期", "date", "time", "数量", "sum", "qty"),
    },
}

_SEMANTIC_ALIASES_DB: dict[str, tuple[str, ...]] = {
    "预测": ("pred", "predict", "forecast", "prob", "result"),
    "实际": ("actual", "real", "qty", "sale", "sales", "sum"),
    "销量": ("qty", "sale", "sales", "sum"),
    "数量": ("qty", "count", "sum", "amount"),
    "准确率": ("accuracy", "acc", "rate", "ratio"),
    "准确": ("accuracy", "acc", "rate"),
    "月份": ("month", "date_month"),
    "日期": ("date", "day", "dt", "time"),
    "时间": ("time", "date", "dt"),
    "国家": ("country", "nation", "region", "area"),
    "地区": ("region", "area", "country"),
    "用户": ("user", "uid", "nickname", "email", "phone"),
    "邮箱": ("email", "mail"),
    "平台": ("platform", "channel"),
}


def _column_semantic_flags(sql_type: str, name: str) -> dict[str, Any]:
    t = (sql_type or "TEXT").upper()
    nm = (name or "").lower()
    is_numeric = any(x in t for x in ("INT", "REAL", "FLOAT", "DOUBLE", "NUM", "DECIMAL", "BOOL"))
    is_time_like = any(x in t for x in ("DATE", "TIME")) or any(
        k in nm
        for k in (
            "时间",
            "日期",
            "time",
            "date",
            "created",
            "updated",
            "注册",
            "活跃",
            "timestamp",
        )
    )
    is_text_like = (
        "CHAR" in t or "TEXT" in t or "CLOB" in t or t == "" or (not is_numeric and not is_time_like)
    )
    low_card_hint = any(
        k in nm for k in ("状态", "类型", "平台", "渠道", "等级", "status", "type", "platform", "category")
    )
    enum_candidate: str
    if is_numeric and "INT" in t:
        enum_candidate = "likely_low" if low_card_hint else "unknown"
    elif is_text_like and low_card_hint:
        enum_candidate = "likely_low"
    elif is_text_like and not low_card_hint:
        enum_candidate = "likely_high"
    else:
        enum_candidate = "unknown"
    return {
        "is_text_like": bool(is_text_like),
        "is_numeric": bool(is_numeric),
        "is_time_like": bool(is_time_like),
        "enum_candidate": enum_candidate,
    }


def _tokenize_for_match(text: str) -> set[str]:
    s = str(text or "").strip().lower()
    raw_latin = re.findall(r"[a-z0-9_]+", s)
    latin: set[str] = set(raw_latin)
    for token in raw_latin:
        latin.update(part for part in re.split(r"_+", token) if part and part != token)
    han: set[str] = set()
    for seq in re.findall(r"[一-鿿]+", s):
        if len(seq) == 1:
            han.add(seq)
            continue
        for n in (2, 3):
            if len(seq) < n:
                continue
            han.update(seq[i : i + n] for i in range(len(seq) - n + 1))
    alias: set[str] = set()
    for trigger, aliases in _SEMANTIC_ALIASES_DB.items():
        if trigger in s:
            alias.update(aliases)
    return latin | han | alias


def _overlap_score_db(question: str, candidate: str) -> int:
    q = str(question or "").strip().lower()
    c = str(candidate or "").strip().lower()
    if not q or not c:
        return 0
    score = 0
    if c in q or q in c:
        score += 8
    score += len(_tokenize_for_match(q) & _tokenize_for_match(c))
    return score


def _normalize_join_key_name(name: str) -> str:
    text = str(name or "").strip().lower()
    if not text:
        return ""
    text = re.sub(r"[`\"'\s\-]+", "", text)
    text = text.replace("_", "")
    text = text.replace("（", "").replace("）", "")
    return text


def _is_id_like_name(name: str) -> bool:
    text = str(name or "")
    low = text.lower()
    return bool(
        text.endswith("ID")
        or text.endswith("编号")
        or "用户ID" in text
        or low == "id"
        or low.endswith("_id")
        or low.endswith("id")
        or "code" in low
    )


def _is_measure_like_name(name: str) -> bool:
    low = str(name or "").strip().lower()
    return any(
        token in low
        for token in (
            "qty",
            "sum",
            "count",
            "avg",
            "mean",
            "price",
            "prob",
            "accuracy",
            "amount",
            "rate",
            "score",
            "数量",
            "合计",
            "金额",
            "概率",
            "准确",
            "均值",
            "平均",
        )
    )


def _join_candidate_score(left: dict[str, Any], right: dict[str, Any]) -> tuple[int, list[str]]:
    ln = str(left.get("name") or "")
    rn = str(right.get("name") or "")
    if not ln or not rn:
        return 0, []
    reasons: list[str] = []
    score = 0
    if ln == rn:
        score += 85
        reasons.append("同名列")
    lnorm = _normalize_join_key_name(ln)
    rnorm = _normalize_join_key_name(rn)
    if lnorm and lnorm == rnorm and ln != rn:
        score += 75
        reasons.append("标准化后同名")
    overlap = _tokenize_for_match(ln) & _tokenize_for_match(rn)
    meaningful_overlap = {tok for tok in overlap if len(str(tok)) >= 2 or any("一" <= ch <= "鿿" for ch in str(tok))}
    if meaningful_overlap:
        score += min(25, 8 * len(meaningful_overlap))
        reasons.append("名称词元重合:" + ", ".join(sorted(map(str, meaningful_overlap))[:3]))
    if left.get("is_time_like") and right.get("is_time_like"):
        score += 22
        reasons.append("两列都像时间键")
    if left.get("is_numeric") and right.get("is_numeric") and _is_id_like_name(ln) and _is_id_like_name(rn):
        score += 25
        reasons.append("两列都像 ID/编号")
    if left.get("is_text_like") and right.get("is_text_like") and meaningful_overlap:
        score += 6
    if _is_measure_like_name(ln) or _is_measure_like_name(rn):
        score -= 40
        reasons.append("更像指标列，已降权")
    if ln == rn and (_is_id_like_name(ln) or left.get("is_time_like") or left.get("enum_candidate") == "likely_low"):
        score += 18
    return max(score, 0), reasons


def _raw_table_column_names(table_name: str) -> set[str]:
    if not is_queryable_table(table_name):
        return set()
    with get_connection() as conn:
        rows = _conn_exec(
            conn,
            """
            SELECT COLUMN_NAME AS name FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
            ORDER BY ORDINAL_POSITION
            """,
            (MYSQL_DATABASE, table_name),
        ).fetchall()
    return {str(r["name"]) for r in rows}


_SQL_SAFE_IGNORE = frozenset(
    """
    select distinct all from where group by order limit offset having as on join inner left right outer
    cross natural union except intersect case when then else end cast is not null like in between exists
    asc desc true false and or count sum avg min max coalesce nullif date time datetime
    curdate current_date current_timestamp now interval true false users
    """.split()
)


def _validate_sql_column_identifiers(normalized: str, lowered: str) -> list[str]:
    tables = {
        token
        for pair in re.findall(r"\bfrom\s+([a-zA-Z_][a-zA-Z0-9_]*)|\bjoin\s+([a-zA-Z_][a-zA-Z0-9_]*)", lowered)
        for token in pair
        if token
    }
    if not tables:
        return []
    all_cols: set[str] = set()
    for t in tables:
        all_cols |= _raw_table_column_names(t)
    idents = re.findall(r"`([^`]+)`", normalized) + re.findall(r'"([^"]+)"', normalized)
    unknown: list[str] = []
    seen: set[str] = set()
    for raw in idents:
        name = raw.strip()
        if not name or name in seen:
            continue
        seen.add(name)
        if name in tables:
            continue
        if name.lower() in _SQL_SAFE_IGNORE:
            continue
        if name not in all_cols:
            unknown.append(name)
    return unknown


def _clean_header(value: Any, index: int) -> str:
    text = str(value or "").strip()
    return text or f"列{index + 1}"


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return value


def _headers_include_id_column(headers: list[str]) -> bool:
    """首行是否已有 id 列（不区分大小写）。有则不可再自动加 INTEGER PRIMARY KEY AUTOINCREMENT 的 id。"""
    return any(str(h or "").strip().lower() == "id" for h in headers)


def _infer_sql_type(values: list[Any]) -> str:
    non_null = [value for value in values if value not in (None, "")]
    if not non_null:
        return "TEXT"
    if all(isinstance(value, bool) for value in non_null):
        return "TINYINT"
    if all(isinstance(value, int) and not isinstance(value, bool) for value in non_null):
        return "BIGINT"
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in non_null):
        return "DOUBLE"
    return "TEXT"


def _read_xlsx(file_path: str | Path) -> tuple[list[str], list[list[Any]]]:
    try:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    except InvalidFileException as exc:
        raise ValueError(
            "无法用 Excel 格式读取（仅支持 .xlsx / .xlsm 等）。若是 CSV，请使用 .csv 扩展名上传。"
        ) from exc
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [_clean_header(value, index) for index, value in enumerate(rows[0])]
    data_rows = [
        [_normalize_cell(row[index]) if index < len(row) else None for index in range(len(headers))]
        for row in rows[1:]
    ]
    return headers, data_rows


def _read_csv(file_path: str | Path) -> tuple[list[str], list[list[Any]]]:
    """首行为表头；尝试 UTF-8（含 BOM）与 GBK；分隔符由 Sniffer 推断（默认可为逗号）。"""
    path = Path(file_path)
    last_decode: Exception | None = None
    rows: list[list[str]] | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            with path.open(newline="", encoding=encoding) as handle:
                sample = handle.read(8192)
                handle.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                except csv.Error:
                    dialect = csv.excel
                rows = list(csv.reader(handle, dialect))
            break
        except UnicodeDecodeError as exc:
            last_decode = exc
            continue
    if rows is None:
        raise ValueError("CSV 无法用 UTF-8 或 GBK 解码，请另存为 UTF-8。") from last_decode
    if not rows:
        return [], []
    headers = [_clean_header(value, index) for index, value in enumerate(rows[0])]
    data_rows = [
        [_normalize_cell(row[index]) if index < len(row) else None for index in range(len(headers))]
        for row in rows[1:]
    ]
    return headers, data_rows


def _read_tabular_file(file_path: str | Path) -> tuple[list[str], list[list[Any]]]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return _read_xlsx(path)
    raise ValueError(
        f"不支持的文件扩展名「{suffix or '无'}」；请使用 .csv 或 .xlsx（以及 .xlsm / .xltx / .xltm）。"
    )


def _migrate_query_logs(conn: pymysql.Connection) -> None:
    rows = _conn_exec(
        conn,
        """
        SELECT COLUMN_NAME AS name FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'query_logs'
        """,
        (MYSQL_DATABASE,),
    ).fetchall()
    columns = {str(row["name"]) for row in rows}
    if "user_feedback" not in columns:
        _conn_exec(conn, "ALTER TABLE `query_logs` ADD COLUMN `user_feedback` INT NULL")


def initialize_database() -> None:
    ensure_directories()
    if not MYSQL_DATABASE:
        return
    with get_connection() as conn:
        _conn_exec(
            conn,
            """
            CREATE TABLE IF NOT EXISTS `query_logs` (
                `id` BIGINT NOT NULL AUTO_INCREMENT PRIMARY KEY,
                `question` TEXT NOT NULL,
                `understanding` TEXT,
                `sql` LONGTEXT,
                `tool_trace` LONGTEXT,
                `created_at` VARCHAR(64) NOT NULL,
                `user_feedback` INT NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """,
        )
        _migrate_query_logs(conn)
        conn.commit()


def import_tabular_to_mysql(file_path: str | Path, target_table: str | None = None) -> tuple[int, list[str], str]:
    """
    用表格首行建列并整表重建后插入全部行。支持 .csv 与 Excel（.xlsx / .xlsm / .xltx / .xltm）。
    target_table 默认 users；可指定其它业务表名（与 INTERNAL_TABLES 不冲突）。
    返回 (行数, 列名列表, 实际表名)。
    """
    table_sql = validate_business_table_name(target_table or USERS_TABLE)
    headers, data_rows = _read_tabular_file(file_path)
    if not headers:
        return 0, [], table_sql

    column_samples = [
        [row[index] for row in data_rows if index < len(row)]
        for index in range(len(headers))
    ]
    column_defs = [
        f"{_quote_identifier(header)} {_infer_sql_type(samples)}"
        for header, samples in zip(headers, column_samples, strict=True)
    ]
    placeholders = ", ".join("%s" for _ in headers)
    quoted_headers = ", ".join(_quote_identifier(header) for header in headers)
    use_synthetic_pk = not _headers_include_id_column(headers)
    body_sql = (
        "`id` BIGINT NOT NULL AUTO_INCREMENT PRIMARY KEY,\n                " + ", ".join(column_defs)
        if use_synthetic_pk
        else ", ".join(column_defs)
    )

    with get_connection() as conn:
        _conn_exec(conn, f"DROP TABLE IF EXISTS {_quote_identifier(table_sql)}")
        _conn_exec(
            conn,
            f"""
            CREATE TABLE {_quote_identifier(table_sql)} (
                {body_sql}
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """,
        )
        if data_rows:
            _conn_executemany(
                conn,
                f"INSERT INTO {_quote_identifier(table_sql)} ({quoted_headers}) VALUES ({placeholders})",
                data_rows,
            )
        conn.commit()

    return len(data_rows), headers, table_sql


def _default_headers() -> list[str]:
    if not DEFAULT_XLSX_PATH.exists():
        return []
    headers, _ = _read_xlsx(DEFAULT_XLSX_PATH)
    return headers


def bootstrap_database() -> None:
    """仅确保 query_logs 存在；业务表由你在 MySQL 中自行维护。"""
    initialize_database()


def count_users() -> int:
    """默认表 `users` 行数；若无该表则用当前库第一张可查询业务表的行数（供 health 展示）。"""
    names = queryable_table_names()
    if not names:
        return 0
    target = next((n for n in names if n.lower() == USERS_TABLE.lower()), names[0])
    return get_table_row_count(target)


def has_imported_data() -> bool:
    return count_users() > 0


def list_tables() -> list[dict[str, Any]]:
    """可查询业务表清单，含行数/列数等便于 agent 路由。"""
    out: list[dict[str, Any]] = []
    for n in queryable_table_names():
        try:
            cols = list_columns(n)
            rc = get_table_row_count(n)
        except pymysql.Error as exc:
            logger.warning("list_tables: 跳过表 %s（读取列或行数失败）: %s", n, exc)
            continue
        out.append(
            {
                "name": n,
                "comment": "",
                "summary": "",
                "row_count": rc,
                "column_count": len(cols),
                "is_default_table": n.lower() == USERS_TABLE.lower(),
            }
        )
    return out


def drop_business_table(table_name: str) -> str:
    """
    删除当前库中的可查询业务表（与 SELECT 白名单一致，不含 query_logs）。
    返回已删除的表名。
    """
    t = validate_business_table_name(str(table_name or "").strip())
    if not _table_exists(t):
        raise ValueError(f"表不存在: {t}")
    allowed = set(queryable_table_names())
    if t not in allowed:
        raise ValueError(f"不允许删除系统表或受保护表: {t}")
    try:
        with get_connection() as conn:
            _conn_exec(conn, f"DROP TABLE {_quote_identifier(t)}")
            conn.commit()
    except pymysql.Error as exc:
        raise ValueError(f"无法删除表（可能被外键引用或权限不足）: {exc}") from exc
    return t


def list_columns(table_name: str = USERS_TABLE) -> list[dict[str, Any]]:
    table = str(table_name or "").strip() or USERS_TABLE
    if not is_queryable_table(table):
        return []
    with get_connection() as conn:
        rows = _conn_exec(
            conn,
            """
            SELECT
                COLUMN_NAME AS name,
                COLUMN_TYPE AS col_type,
                IS_NULLABLE AS is_nullable,
                COLUMN_KEY AS col_key,
                COLUMN_COMMENT AS col_comment
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
            ORDER BY ORDINAL_POSITION
            """,
            (MYSQL_DATABASE, table),
        ).fetchall()
    columns: list[dict[str, Any]] = []
    for row in rows:
        name = row["name"]
        sql_type = str(row["col_type"] or "TEXT")
        flags = _column_semantic_flags(sql_type, name)
        nullable = str(row.get("is_nullable") or "YES").upper() == "YES"
        pk = str(row.get("col_key") or "") == "PRI"
        comment = str(row.get("col_comment") or "").strip()
        columns.append(
            {
                "name": name,
                "label": comment or name,
                "sql_type": sql_type,
                "nullable": nullable,
                "primary_key": pk,
                "comment": comment,
                **flags,
            }
        )
    return columns


def get_table_row_count(table_name: str) -> int:
    if not is_queryable_table(table_name):
        return 0
    with get_connection() as conn:
        row = _conn_exec(conn, f"SELECT COUNT(*) AS n FROM {_quote_identifier(table_name)}").fetchone()
    return int(row["n"])


def get_table_schema(table_name: str = USERS_TABLE) -> dict[str, Any]:
    table = str(table_name or "").strip() or USERS_TABLE
    if not is_queryable_table(table):
        return {"table": table, "row_count": 0, "columns": []}
    return {"table": table, "row_count": get_table_row_count(table), "columns": list_columns(table)}


def get_table_relationships() -> list[dict[str, Any]]:
    """基于 information_schema 外键；未建 FK 的库返回空列表。"""
    relationships: list[dict[str, Any]] = []
    with get_connection() as conn:
        rows = _conn_exec(
            conn,
            """
            SELECT TABLE_NAME AS tbl, COLUMN_NAME AS col,
                   REFERENCED_TABLE_NAME AS ref_tbl, REFERENCED_COLUMN_NAME AS ref_col
            FROM information_schema.KEY_COLUMN_USAGE
            WHERE TABLE_SCHEMA = %s
              AND REFERENCED_TABLE_SCHEMA = %s
              AND REFERENCED_TABLE_NAME IS NOT NULL
            """,
            (MYSQL_DATABASE, MYSQL_DATABASE),
        ).fetchall()
    allowed = set(queryable_table_names())
    for fk in rows:
        ft = str(fk["tbl"])
        tt = str(fk["ref_tbl"])
        if ft not in allowed or tt not in allowed:
            continue
        relationships.append(
            {
                "from_table": ft,
                "from_column": str(fk["col"]),
                "to_table": tt,
                "to_column": str(fk["ref_col"]),
                "relationship": "many_to_one",
            }
        )
    return relationships


def infer_join_candidates(tables: list[str] | None = None, limit: int = 20) -> dict[str, Any]:
    """
    在没有显式外键时，基于列名/语义给出可疑的等值 JOIN 候选。
    目标不是 100% 自动正确，而是给 Agent 一个比“写死 users 表”更泛化的跨表探查入口。
    """
    selected = []
    allowed = set(queryable_table_names())
    for raw in tables or queryable_table_names():
        name = str(raw or "").strip()
        if name and name in allowed and name not in selected:
            selected.append(name)
    if len(selected) < 2:
        return {
            "ok": True,
            "tables": selected,
            "candidates": [],
            "meta": {"message": "少于两张表，暂无 JOIN 候选"},
        }

    limit = min(max(int(limit), 1), 50)
    schema_map = {table: list_columns(table) for table in selected}
    candidates: list[dict[str, Any]] = []
    for i, left_table in enumerate(selected):
        for right_table in selected[i + 1 :]:
            left_cols = schema_map.get(left_table, [])
            right_cols = schema_map.get(right_table, [])
            for left in left_cols:
                for right in right_cols:
                    score, reasons = _join_candidate_score(left, right)
                    if score < 60:
                        continue
                    left_name = str(left.get("name") or "")
                    right_name = str(right.get("name") or "")
                    candidates.append(
                        {
                            "from_table": left_table,
                            "from_column": left_name,
                            "to_table": right_table,
                            "to_column": right_name,
                            "on": f'{_quote_identifier(left_table)}.{_quote_identifier(left_name)} = {_quote_identifier(right_table)}.{_quote_identifier(right_name)}',
                            "score": score,
                            "confidence": "high" if score >= 95 else "medium",
                            "reasons": reasons,
                        }
                    )
    candidates.sort(key=lambda item: (item["score"], item["from_table"], item["to_table"]), reverse=True)

    deduped: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for item in candidates:
        key = (
            item["from_table"],
            item["from_column"],
            item["to_table"],
            item["to_column"],
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
        if len(deduped) >= limit:
            break
    return {
        "ok": True,
        "tables": selected,
        "candidates": deduped,
        "meta": {"candidate_count": len(deduped), "tables_considered": len(selected)},
    }


_ISO_DATETIME_TEXT_RE = re.compile(
    r"^\s*\d{4}-\d{2}-\d{2}(?:[ T]\d{2}:\d{2}(?::\d{2})?)?\s*$",
    re.ASCII,
)


def _inspect_default_order_sql(table_name: str, available_columns: list[str]) -> str:
    meta_cols = list_columns(table_name)
    for c in meta_cols:
        nm = str(c.get("name") or "")
        if c.get("primary_key") and nm in available_columns:
            return f" ORDER BY {_quote_identifier(nm)} DESC"
    if available_columns:
        return f" ORDER BY {_quote_identifier(available_columns[0])} DESC"
    return ""


def inspect_rows(
    table: str = USERS_TABLE,
    limit: int = 5,
    filters: dict[str, Any] | None = None,
    as_dict: bool = True,
    order_by: str | None = None,
    columns: list[str] | None = None,
) -> dict[str, Any]:
    """统一表样本探查：默认每行为列名→值的字典。可选 ``columns`` 只返回指定列（须在表中存在）。"""
    table_name = resolve_table_name(table)
    if not is_queryable_table(table_name):
        return {
            "ok": False,
            "error": f"Unknown or disallowed table: {table_name}",
            "table": table_name,
            "columns": [],
            "rows": [],
            "meta": {},
        }
    filters = filters or {}
    safe_limit = min(max(limit, 1), 50)
    available_columns = column_names_for_table(table_name)
    if not available_columns:
        return {
            "ok": True,
            "error": None,
            "table": table_name,
            "columns": [],
            "rows": [],
            "meta": {"limit": safe_limit},
        }

    requested = [str(c).strip() for c in (columns or []) if str(c).strip()]
    invalid_requested = [c for c in requested if c not in available_columns]
    if requested:
        select_list = [c for c in requested if c in available_columns]
        if not select_list:
            select_list = list(available_columns)
            invalid_requested = requested
    else:
        select_list = list(available_columns)

    clauses: list[str] = []
    params: list[Any] = []
    for field, value in filters.items():
        if field not in available_columns:
            continue
        clauses.append(f"{_quote_identifier(field)} = %s")
        params.append(value)

    select_columns = ", ".join(_quote_identifier(column) for column in select_list)
    qtable = _quote_identifier(table_name)
    sql = f"SELECT {select_columns} FROM {qtable}"
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    ob = str(order_by or "").strip()
    order_col = ob.split(".")[-1] if ob else ""
    if order_col in available_columns:
        sql += f" ORDER BY {_quote_identifier(order_col)} DESC"
    else:
        sql += _inspect_default_order_sql(table_name, select_list)
    sql += " LIMIT %s"
    params.append(safe_limit)
    meta: dict[str, Any] = {"limit": safe_limit, "as_dict": as_dict}
    if invalid_requested:
        meta["columns_ignored_unknown"] = invalid_requested
    if requested and select_list != available_columns:
        meta["columns_selected"] = select_list
    with get_connection() as conn:
        rows = _conn_exec(conn, sql, tuple(params)).fetchall()
    if as_dict:
        dict_rows = [{c: row[c] for c in select_list} for row in rows]
        return {
            "ok": True,
            "error": None,
            "table": table_name,
            "columns": select_list,
            "rows": dict_rows,
            "meta": {**meta, "as_dict": True},
        }
    return {
        "ok": True,
        "error": None,
        "table": table_name,
        "columns": select_list,
        "rows": [_row_to_ordered_values(row, select_list) for row in rows],
        "meta": {**meta, "as_dict": False},
    }


def profile_time_column(
    table: str = USERS_TABLE,
    column: str | None = None,
    field: str | None = None,
    sample_limit: int = 12,
) -> dict[str, Any]:
    """
    为「按注册/创建/活跃日期区间筛选」提供**数据驱动**的统计：非空 min/max、样例值、
    是否像 ISO 日期时间字符串，以及**通用** MySQL WHERE 写法说明（不把具体业务日期写死在代码里）。
    """
    table_name = resolve_table_name(table)
    if not is_queryable_table(table_name):
        return {
            "ok": False,
            "error": f"Unknown or disallowed table: {table_name}",
            "table": table_name,
            "column": column or field,
            "min_value": None,
            "max_value": None,
            "sample_values": [],
            "looks_iso_datetime_text": False,
            "is_time_like_by_schema": False,
            "filter_hint": "",
        }
    cname = str(column or field or "").strip()
    if not cname:
        return {
            "ok": False,
            "error": "必须提供 column（或 field）为表中列名。",
            "table": table_name,
            "column": None,
            "min_value": None,
            "max_value": None,
            "sample_values": [],
            "looks_iso_datetime_text": False,
            "is_time_like_by_schema": False,
            "filter_hint": "",
        }
    available = column_names_for_table(table_name)
    if cname not in available:
        return {
            "ok": False,
            "error": f"列 `{cname}` 不在表 `{table_name}` 中。可用列: {', '.join(available[:40])}"
            + ("…" if len(available) > 40 else ""),
            "table": table_name,
            "column": cname,
            "min_value": None,
            "max_value": None,
            "sample_values": [],
            "looks_iso_datetime_text": False,
            "is_time_like_by_schema": False,
            "filter_hint": "",
        }
    col_meta = next((x for x in list_columns(table_name) if x.get("name") == cname), {})
    is_time_like = bool(col_meta.get("is_time_like"))
    qc = _quote_identifier(cname)
    qt = _quote_identifier(table_name)
    sl = min(max(int(sample_limit or 12), 4), 40)
    with get_connection() as conn:
        agg = _conn_exec(conn,
            f"""
            SELECT
              MIN(TRIM(CAST({qc} AS CHAR))) AS vmin,
              MAX(TRIM(CAST({qc} AS CHAR))) AS vmax
            FROM {qt}
            WHERE {qc} IS NOT NULL AND TRIM(CAST({qc} AS CHAR)) != ''
            """
        ).fetchone()
        vmin = agg["vmin"] if agg else None
        vmax = agg["vmax"] if agg else None
        low = [
            _dict_cursor_cell(r, "v")
            for r in _conn_exec(conn,
                f"""
                SELECT TRIM(CAST({qc} AS CHAR)) AS v FROM {qt}
                WHERE {qc} IS NOT NULL AND TRIM(CAST({qc} AS CHAR)) != ''
                ORDER BY v ASC LIMIT 3
                """
            ).fetchall()
        ]
        high = [
            _dict_cursor_cell(r, "v")
            for r in _conn_exec(conn,
                f"""
                SELECT TRIM(CAST({qc} AS CHAR)) AS v FROM {qt}
                WHERE {qc} IS NOT NULL AND TRIM(CAST({qc} AS CHAR)) != ''
                ORDER BY v DESC LIMIT 3
                """
            ).fetchall()
        ]
    samples: list[str] = []
    seen: set[str] = set()
    for v in low + high:
        if v is None:
            continue
        s = str(v)
        if s not in seen:
            seen.add(s)
            samples.append(s)
        if len(samples) >= sl:
            break
    iso_hits = sum(1 for s in samples if _ISO_DATETIME_TEXT_RE.match(s))
    looks_iso = len(samples) > 0 and iso_hits >= max(1, int(0.7 * len(samples)))
    hint_parts = [
        "根据当前表该列**实际**非空 min/max 与样例推断格式；把问题里的起止日期代入下列模式即可，勿虚构列名。",
    ]
    if looks_iso or is_time_like:
        hint_parts.append(
            f"若取值形如 `YYYY-MM-DD` 或 `YYYY-MM-DD HH:MM:SS`，同一格式的字符串**字典序与日历序一致**。"
            f"含两端整日的闭区间 `[开始日, 结束日]` 常用写法："
            f"`` `{{表}}`.`{cname}` >= '开始日 00:00:00' AND `{{表}}`.`{cname}` < '结束日+1天 00:00:00'``（无时分秒时可只用日期串比较边界）；"
            f"或 ``DATE(`{{表}}`.`{cname}`) BETWEEN '开始日' AND '结束日'``（列已为 DATE/DATETIME 时）。"
        )
    else:
        hint_parts.append(
            "样例**不像**统一 ISO 日期串：请结合 Inspect_rows 的 `columns` 只看该列，或 Profile_column 看高频值，再选 DATE()/STR_TO_DATE、LIKE 或子串比较。"
        )
    hint_parts.append(
        "另：「有邮箱」类条件应对**邮箱列**使用 "
        "``(`邮箱列` IS NOT NULL AND TRIM(CAST(`邮箱列` AS CHAR)) <> '')``（列名以 schema 为准）。"
    )
    return {
        "ok": True,
        "error": None,
        "table": table_name,
        "column": cname,
        "sql_type": col_meta.get("sql_type"),
        "is_time_like_by_schema": is_time_like,
        "min_value": vmin,
        "max_value": vmax,
        "sample_values": samples,
        "looks_iso_datetime_text": looks_iso,
        "filter_hint": " ".join(hint_parts),
    }


def profile_column(
    table: str = USERS_TABLE,
    column: str | None = None,
    field: str | None = None,
    include_top_values: bool = True,
    top_k: int = 10,
    keyword: str | None = None,
    sample_cap: int = 40,
) -> dict[str, Any]:
    """单列画像：合并原 field_profile + distinct_values（含 keyword 过滤，大小写不敏感）。"""
    table_name = resolve_table_name(table)
    col = str(column or field or "").strip()
    if not col:
        return {"ok": False, "error": "column/field 不能为空", "table": table_name, "data": None}
    names = set(column_names_for_table(table_name))
    if col not in names:
        return {"ok": False, "error": f"Unknown field: {col}", "table": table_name, "data": None}
    top_k = min(max(int(top_k), 1), 50)
    sample_cap = min(max(int(sample_cap), 5), 80)
    quoted_field = _quote_identifier(col)
    qtable = _quote_identifier(table_name)
    clauses = [f"{quoted_field} IS NOT NULL", f"TRIM(CAST({quoted_field} AS CHAR)) != ''"]
    params: list[Any] = []
    if keyword and str(keyword).strip():
        clauses.append("INSTR(LOWER(CAST({0} AS CHAR)), %s) > 0".format(quoted_field))
        params.append(str(keyword).strip().lower())
    where_sql = " AND ".join(clauses)
    with get_connection() as conn:
        stat = _conn_exec(conn,
            f"""
            SELECT COUNT(*) AS total_rows,
                   COUNT({quoted_field}) AS non_null_rows,
                   COUNT(DISTINCT {quoted_field}) AS distinct_count
            FROM {qtable}
            """
        ).fetchone()
        top_values: list[dict[str, Any]] = []
        if include_top_values:
            top_values = [
                dict(row)
                for row in _conn_exec(conn,
                    f"""
                    SELECT {quoted_field} AS value, COUNT(*) AS count
                    FROM {qtable}
                    WHERE {where_sql}
                    GROUP BY {quoted_field}
                    ORDER BY count DESC, value ASC
                    LIMIT %s
                    """,
                    [*params, top_k],
                ).fetchall()
            ]
        sample_rows = _conn_exec(
            conn,
            f"""
            SELECT {quoted_field}
            FROM {qtable}
            WHERE {where_sql}
            GROUP BY {quoted_field}
            ORDER BY COUNT(*) DESC
            LIMIT %s
            """,
            [*params, sample_cap],
        ).fetchall()
        samples = [_dict_cursor_cell(row, col) for row in sample_rows]
    total = int(stat["total_rows"])
    non_null = int(stat["non_null_rows"])
    distinct = int(stat["distinct_count"])
    null_ratio = (total - non_null) / total if total else 0.0
    distinct_ratio = distinct / non_null if non_null else 0.0
    data = {
        "column": col,
        "total_rows": total,
        "non_null_rows": non_null,
        "null_ratio": round(null_ratio, 6),
        "distinct_count": distinct,
        "distinct_ratio": round(distinct_ratio, 6),
        "top_values": top_values,
        "sample_values": samples,
    }
    return {"ok": True, "error": None, "table": table_name, "data": data}


def preview_rows(
    limit: int = 5,
    filters: dict[str, Any] | None = None,
    table: str = USERS_TABLE,
) -> dict[str, Any]:
    """兼容旧工具：返回 columns + 行数组（limit 上限 20 与历史行为一致）。"""
    lim = min(max(limit, 1), 20)
    r = inspect_rows(table=table, limit=lim, filters=filters, as_dict=False)
    if not r.get("ok"):
        return {"table": r.get("table", table), "columns": [], "rows": []}
    return {"table": r["table"], "columns": r["columns"], "rows": r["rows"]}


def field_profile(field: str, table: str = USERS_TABLE) -> dict[str, Any]:
    """兼容：映射到 profile_column。"""
    r = profile_column(table=table, field=field, include_top_values=True, top_k=50, sample_cap=40)
    if not r.get("ok"):
        return {"error": r.get("error"), "table": r.get("table")}
    d = r["data"] or {}
    return {
        "table": r["table"],
        "field": field,
        "total_rows": d.get("total_rows"),
        "non_null_rows": d.get("non_null_rows"),
        "distinct_count": d.get("distinct_count"),
        "sample_values": d.get("sample_values", []),
    }


def distinct_values(
    field: str,
    keyword: str | None = None,
    limit: int = 10,
    table: str = USERS_TABLE,
) -> dict[str, Any]:
    """兼容：映射到 profile_column 的 top_values。"""
    r = profile_column(
        table=table,
        field=field,
        keyword=keyword,
        include_top_values=True,
        top_k=limit,
        sample_cap=30,
    )
    if not r.get("ok"):
        return {"error": r.get("error"), "table": r.get("table"), "field": field, "values": []}
    vals = (r.get("data") or {}).get("top_values", [])
    return {"table": r["table"], "field": field, "values": vals}


def _dedupe_preserve_text(items: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for raw in items:
        value = str(raw or '').strip()
        if not value:
            continue
        key = value.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(value)
    return out


def _infer_filter_semantic_type(
    query: str | None = None,
    value: str | None = None,
    semantic_type: str | None = None,
) -> str:
    explicit = str(semantic_type or '').strip().lower()
    if explicit in {'country', 'platform', 'category', 'time', 'id', 'generic'}:
        return explicit
    blob = f"{query or ''} {value or ''}".lower()
    if any(tok in blob for tok in ('国家', '地区', '区域', '国别', 'country', 'nation', 'region', 'market', '市场', '城市', 'city')):
        return 'country'
    if any(tok in blob for tok in ('平台', '渠道', '来源', 'platform', 'channel', 'device', 'terminal', '终端', 'os')):
        return 'platform'
    if any(tok in blob for tok in ('状态', '类型', '身份', '角色', '标签', 'status', 'type', 'role', 'identity', 'tag', 'category')):
        return 'category'
    if any(tok in blob for tok in ('时间', '日期', '活跃', '注册', '创建', 'date', 'time', 'month', 'day', 'year')):
        return 'time'
    if any(tok in blob for tok in ('用户id', '订单号', '客户id', 'matnr', 'sku', 'uid', '编号', '编码', '单号', ' code', 'id')):
        return 'id'
    return 'generic'


def _expand_keyword_variants(
    keyword_variants: list[str] | None = None,
    keyword: str | None = None,
    value: str | None = None,
    query: str | None = None,
    semantic_type: str | None = None,
) -> list[str]:
    seeds: list[str] = []
    if isinstance(keyword_variants, list):
        seeds.extend(str(x or '').strip() for x in keyword_variants if str(x or '').strip())
    for raw in (keyword, value):
        if str(raw or '').strip():
            seeds.append(str(raw).strip())
    joined = ' '.join(seeds + [str(query or '').strip()]).strip()
    resolved_semantic = _infer_filter_semantic_type(query=query, value=value or keyword, semantic_type=semantic_type)
    if resolved_semantic == 'country' and joined:
        intent = resolve_country_intent(joined)
        if intent:
            seeds.extend([intent.cn_display, intent.canonical_en, *intent.triggers])
            for phone_prefix in _COUNTRY_PHONE_PREFIX_HINTS.get(intent.canonical_en.strip().lower(), tuple()):
                seeds.append(phone_prefix)
    return _dedupe_preserve_text(seeds)




def _normalize_value_text(value: Any) -> str:
    text = str(value or '').strip().lower()
    text = re.sub(r"[\s\-_/\.,;:()\[\]{}]+", "", text)
    return text


def _value_match_score(raw_value: str, candidate: str, variants: list[str]) -> tuple[float, list[str]]:
    base = _normalize_value_text(raw_value)
    cand = _normalize_value_text(candidate)
    reasons: list[str] = []
    if not cand:
        return 0.0, reasons
    score = 0.0
    if base and cand == base:
        score += 1.0
        reasons.append('与原始值归一化后完全一致')
    for variant in variants:
        v = _normalize_value_text(variant)
        if not v:
            continue
        if cand == v:
            score = max(score, 0.99)
            reasons.append(f'与候选别名 {variant} 完全一致')
        elif v and (v in cand or cand in v):
            score = max(score, 0.88)
            reasons.append(f'与候选别名 {variant} 存在子串对齐')
    ratio = difflib.SequenceMatcher(None, base or raw_value.lower(), cand).ratio() if (base or raw_value) and cand else 0.0
    if ratio > score:
        score = ratio
        reasons.append(f'归一化相似度 {ratio:.2f}')
    return min(score, 1.0), reasons


def _candidate_literals_for_value(raw_value: str, semantic_type: str | None = None, query: str | None = None, candidate_literals: list[str] | None = None) -> list[str]:
    seeds: list[str] = []
    if isinstance(candidate_literals, list):
        seeds.extend(str(x or '').strip() for x in candidate_literals if str(x or '').strip())
    if str(raw_value or '').strip():
        seeds.append(str(raw_value).strip())
    context = ' '.join(x for x in [str(query or '').strip(), str(raw_value or '').strip()] if x)
    resolved = _infer_filter_semantic_type(query=query, value=raw_value, semantic_type=semantic_type)
    if resolved == 'country' and context:
        seeds.extend(country_literal_candidates(context))
    return _dedupe_preserve_text(seeds)


def _fetch_distinct_text_values(table_name: str, column_name: str, limit: int = 50) -> list[dict[str, Any]]:
    qtable = _quote_identifier(table_name)
    qcol = _quote_identifier(column_name)
    safe_limit = min(max(int(limit), 1), 200)
    with get_connection() as conn:
        rows = _conn_exec(conn,
            f"""
            SELECT CAST({qcol} AS CHAR) AS value, COUNT(*) AS cnt
            FROM {qtable}
            WHERE {qcol} IS NOT NULL AND TRIM(CAST({qcol} AS CHAR)) != ''
            GROUP BY CAST({qcol} AS CHAR)
            ORDER BY cnt DESC, value ASC
            LIMIT %s
            """,
            (safe_limit,),
        ).fetchall()
    return [{"value": row['value'], "count": int(row['cnt'])} for row in rows if row['value'] is not None]


def resolve_filter_value(
    raw_value: str,
    semantic_type: str | None = None,
    table: str | None = None,
    column: str | None = None,
    query: str | None = None,
    candidate_literals: list[str] | None = None,
    limit: int = 50,
) -> dict[str, Any]:
    """把用户过滤值归一化为可用于 SQL 的候选字面值；如给定列，则结合列真实值做对齐。"""
    value = str(raw_value or '').strip()
    if not value:
        return {"ok": False, "error": 'raw_value 不能为空', "raw_value": raw_value}
    resolved_semantic = _infer_filter_semantic_type(query=query, value=value, semantic_type=semantic_type)
    variants = _candidate_literals_for_value(value, semantic_type=resolved_semantic, query=query, candidate_literals=candidate_literals)
    out: dict[str, Any] = {
        "ok": True,
        "error": None,
        "raw_value": value,
        "semantic_type": resolved_semantic,
        "candidate_literals": variants,
        "normalized_value": variants[1] if len(variants) > 1 else variants[0],
        "display_value": variants[0],
        "match_strategy": 'alias_dictionary' if len(variants) > 1 else 'identity',
        "confidence": 'high' if len(variants) > 1 else 'medium',
    }
    if str(table or '').strip() and str(column or '').strip():
        aligned = match_value_in_column(
            table=str(table),
            column=str(column),
            raw_value=value,
            semantic_type=resolved_semantic,
            candidate_literals=variants,
            limit=limit,
        )
        out['column_alignment'] = aligned
        if aligned.get('ok') and aligned.get('recommended_sql_literal'):
            out['recommended_sql_literal'] = aligned.get('recommended_sql_literal')
            out['confidence'] = aligned.get('confidence', out['confidence'])
            out['match_strategy'] = aligned.get('strategy', out['match_strategy'])
    return out


def match_value_in_column(
    table: str,
    column: str,
    raw_value: str,
    semantic_type: str | None = None,
    candidate_literals: list[str] | None = None,
    limit: int = 50,
) -> dict[str, Any]:
    """结合目标列真实 distinct values，把用户值对齐为库内更可能存在的字面值。"""
    table_name = resolve_table_name(table)
    if not is_queryable_table(table_name):
        return {"ok": False, "error": f"Unknown or disallowed table: {table_name}", "table": table_name, "column": column}
    column_name = str(column or '').strip()
    if not column_name:
        return {"ok": False, "error": 'column 不能为空', "table": table_name, "column": column_name}
    if column_name not in column_names_for_table(table_name):
        return {"ok": False, "error": f"Unknown column: {column_name}", "table": table_name, "column": column_name}
    value = str(raw_value or '').strip()
    if not value:
        return {"ok": False, "error": 'raw_value 不能为空', "table": table_name, "column": column_name}
    resolved_semantic = _infer_filter_semantic_type(value=value, semantic_type=semantic_type)
    variants = _candidate_literals_for_value(value, semantic_type=resolved_semantic, candidate_literals=candidate_literals)
    distinct_values = _fetch_distinct_text_values(table_name, column_name, limit=limit)
    matches: list[dict[str, Any]] = []
    for item in distinct_values:
        cand = str(item.get('value') or '').strip()
        score, reasons = _value_match_score(value, cand, variants)
        if score < 0.45:
            continue
        matches.append({
            'value': cand,
            'count': int(item.get('count') or 0),
            'score': round(score, 4),
            'reason': '; '.join(reasons[:3]) or '值相似',
        })
    matches.sort(key=lambda row: (float(row.get('score') or 0.0), int(row.get('count') or 0), str(row.get('value') or '')), reverse=True)
    best = matches[0] if matches else None
    if best and float(best.get('score') or 0.0) >= 0.92:
        confidence = 'high'
    elif best and float(best.get('score') or 0.0) >= 0.7:
        confidence = 'medium'
    else:
        confidence = 'low'
    return {
        'ok': True,
        'error': None,
        'table': table_name,
        'column': column_name,
        'raw_value': value,
        'semantic_type': resolved_semantic,
        'candidate_literals': variants,
        'matched_values': matches[:10],
        'recommended_sql_literal': best.get('value') if best else None,
        'strategy': 'column_value_alignment' if best else 'no_column_match',
        'confidence': confidence,
    }


def build_value_predicate(
    table: str,
    column: str,
    raw_value: str,
    semantic_type: str | None = None,
    query: str | None = None,
    candidate_literals: list[str] | None = None,
    prefer_exact: bool = True,
    limit: int = 50,
) -> dict[str, Any]:
    """根据归一化值 + 列值对齐结果生成更稳的 SQL 谓词。"""
    table_name = resolve_table_name(table)
    column_name = str(column or '').strip()
    if not is_queryable_table(table_name):
        return {"ok": False, "error": f"Unknown or disallowed table: {table_name}", "table": table_name, "column": column_name}
    if column_name not in column_names_for_table(table_name):
        return {"ok": False, "error": f"Unknown column: {column_name}", "table": table_name, "column": column_name}
    resolved = resolve_filter_value(
        raw_value=raw_value,
        semantic_type=semantic_type,
        table=table_name,
        column=column_name,
        query=query,
        candidate_literals=candidate_literals,
        limit=limit,
    )
    if not resolved.get('ok'):
        return resolved
    alignment = resolved.get('column_alignment') or {}
    literal = str(alignment.get('recommended_sql_literal') or '').strip()
    variants = list(resolved.get('candidate_literals') or [])
    col_meta = next((c for c in list_columns(table_name) if str(c.get('name') or '') == column_name), {})
    is_text_like = bool(col_meta.get('is_text_like'))
    enum_like = str(col_meta.get('enum_candidate') or '') == 'likely_low'
    best_score = 0.0
    if isinstance(alignment.get('matched_values'), list) and alignment.get('matched_values'):
        try:
            best_score = float((alignment.get('matched_values') or [{}])[0].get('score') or 0.0)
        except (TypeError, ValueError):
            best_score = 0.0
    qcol = f"`{table_name}`.`{column_name}`"
    should_exact = bool(literal) and (not is_text_like or enum_like or (prefer_exact and best_score >= 0.96))
    if should_exact:
        escaped = literal.replace("'", "''")
        predicate = f"{qcol} = '{escaped}'"
        strategy = 'exact_match_on_aligned_value'
    elif literal and is_text_like:
        escaped = literal.lower().replace("'", "''")
        predicate = f"LOWER(CAST({qcol} AS CHAR)) LIKE '%{escaped}%'"
        strategy = 'like_on_aligned_value'
    elif variants:
        clauses: list[str] = []
        low_expr = f"LOWER(CAST({qcol} AS CHAR))"
        for variant in variants[:8]:
            v = str(variant or '').strip()
            if not v:
                continue
            escaped = v.lower().replace("'", "''")
            clauses.append(f"{low_expr} LIKE '%{escaped}%'")
        predicate = '(' + ' OR '.join(clauses) + ')' if clauses else '1=0'
        strategy = 'text_fallback'
    else:
        predicate = '1=0'
        strategy = 'no_viable_literal'
    return {
        'ok': True,
        'error': None,
        'table': table_name,
        'column': column_name,
        'raw_value': str(raw_value or '').strip(),
        'semantic_type': resolved.get('semantic_type'),
        'candidate_literals': variants,
        'recommended_sql_literal': literal or None,
        'sql_predicate': predicate,
        'strategy': strategy,
        'value_resolution': resolved,
    }

def preview_distinct_values(
    table: str = USERS_TABLE,
    column: str | None = None,
    field: str | None = None,
    keyword: str | None = None,
    limit: int = 10,
) -> dict[str, Any]:
    """显式值域预览：返回单列 top distinct values + 样例，用于判断列是否承载国家/平台/状态等维度。"""
    table_name = resolve_table_name(table)
    if not is_queryable_table(table_name):
        return {"ok": False, "error": f"Unknown or disallowed table: {table_name}", "table": table_name, "column": column or field, "values": []}
    cname = str(column or field or '').strip()
    if not cname:
        return {"ok": False, "error": '必须提供 column（或 field）', "table": table_name, "column": None, "values": []}
    if cname not in column_names_for_table(table_name):
        return {"ok": False, "error": f"Unknown column: {cname}", "table": table_name, "column": cname, "values": []}
    top_k = min(max(int(limit), 1), 30)
    prof = profile_column(
        table=table_name,
        column=cname,
        keyword=keyword,
        include_top_values=True,
        top_k=top_k,
        sample_cap=min(max(top_k, 5), 20),
    )
    if not prof.get('ok'):
        return {"ok": False, "error": prof.get('error'), "table": table_name, "column": cname, "values": []}
    data = prof.get('data') or {}
    return {
        "ok": True,
        "error": None,
        "table": table_name,
        "column": cname,
        "keyword": str(keyword or '').strip() or None,
        "values": list(data.get('top_values', []) or []),
        "sample_values": list(data.get('sample_values', []) or []),
        "distinct_count": int(data.get('distinct_count') or 0),
        "non_null_rows": int(data.get('non_null_rows') or 0),
    }


def _search_value_examples_in_column(
    table_name: str,
    column: str,
    variants: list[str],
    limit_per_column: int,
) -> dict[str, Any]:
    qtable = _quote_identifier(table_name)
    qcol = _quote_identifier(column)
    results: list[dict[str, Any]] = []
    total_match_count = 0
    with get_connection() as conn:
        for variant in variants:
            needle = str(variant or '').strip().lower()
            if not needle:
                continue
            count_row = _conn_exec(conn,
                f"""
                SELECT COUNT(*) AS cnt
                FROM {qtable}
                WHERE {qcol} IS NOT NULL
                  AND TRIM(CAST({qcol} AS CHAR)) != ''
                  AND INSTR(LOWER(CAST({qcol} AS CHAR)), %s) > 0
                """,
                (needle,),
            ).fetchone()
            cnt = int(count_row['cnt']) if count_row else 0
            if cnt <= 0:
                continue
            total_match_count += cnt
            rows = _conn_exec(conn,
                f"""
                SELECT CAST({qcol} AS CHAR) AS value, COUNT(*) AS cnt
                FROM {qtable}
                WHERE {qcol} IS NOT NULL
                  AND TRIM(CAST({qcol} AS CHAR)) != ''
                  AND INSTR(LOWER(CAST({qcol} AS CHAR)), %s) > 0
                GROUP BY CAST({qcol} AS CHAR)
                ORDER BY cnt DESC, value ASC
                LIMIT %s
                """,
                (needle, limit_per_column),
            ).fetchall()
            results.append(
                {
                    'variant': variant,
                    'match_count': cnt,
                    'examples': [
                        {'value': row['value'], 'count': int(row['cnt'])}
                        for row in rows
                        if row['value'] is not None
                    ],
                }
            )
    return {
        'column': column,
        'total_match_count': total_match_count,
        'matched_variants': [item['variant'] for item in results],
        'variant_hits': results,
    }


def search_value_examples(
    keyword_variants: list[str] | None = None,
    keyword: str | None = None,
    value: str | None = None,
    table: str = USERS_TABLE,
    columns: list[str] | None = None,
    semantic_type: str | None = None,
    limit_per_column: int = 3,
    max_columns: int = 12,
) -> dict[str, Any]:
    """在候选列里同时搜索多种值写法，并返回样例，适合做国家/平台/标签等 filter grounding。"""
    table_name = resolve_table_name(table)
    if not is_queryable_table(table_name):
        return {"ok": False, "error": f"Unknown or disallowed table: {table_name}", "table": table_name, "columns": []}
    resolved_semantic = _infer_filter_semantic_type(query=keyword, value=value or keyword, semantic_type=semantic_type)
    variants = _expand_keyword_variants(
        keyword_variants=keyword_variants,
        keyword=keyword,
        value=value,
        query=keyword,
        semantic_type=resolved_semantic,
    )
    if not variants:
        return {"ok": False, "error": '必须提供 keyword / value / keyword_variants 之一', "table": table_name, "columns": []}

    available = column_names_for_table(table_name)
    requested = [str(x or '').strip() for x in (columns or []) if str(x or '').strip()]
    if requested:
        targets = [name for name in requested if name in available]
        selection_source = 'user_columns'
    else:
        cols_meta = list_columns(table_name)
        probe_seed = ' '.join(variants)
        targets, probe_meta = _keyword_probe_target_columns(cols_meta, probe_seed, None)
        selection_source = str(probe_meta.get('scope') or 'auto')
    if not requested:
        targets = targets[: min(max(int(max_columns), 1), 24)]
    if not targets:
        return {"ok": False, "error": '没有可搜索的候选列', "table": table_name, "columns": []}

    safe_limit = min(max(int(limit_per_column), 1), 10)
    out: list[dict[str, Any]] = []
    for column_name in targets:
        item = _search_value_examples_in_column(table_name, column_name, variants, safe_limit)
        example_values = []
        seen_values: set[str] = set()
        for hit in item.get('variant_hits', []):
            for ex in hit.get('examples', []):
                v = str(ex.get('value') or '')
                if not v or v in seen_values:
                    continue
                seen_values.add(v)
                example_values.append(v)
        item['example_values'] = example_values[:safe_limit]
        item['confidence'] = 'high' if item['total_match_count'] >= 5 else ('medium' if item['total_match_count'] > 0 else 'low')
        out.append(item)
    out.sort(key=lambda row: (int(row.get('total_match_count') or 0), row.get('column') or ''), reverse=True)
    return {
        "ok": True,
        "error": None,
        "table": table_name,
        "semantic_type": resolved_semantic,
        "variants": variants,
        "columns": out,
        "meta": {
            'searched_columns': targets,
            'selection_source': selection_source,
            'searched_variants': len(variants),
        },
    }


def _column_semantic_match_score(
    column_meta: dict[str, Any],
    query: str,
    semantic_type: str,
) -> tuple[int, list[str]]:
    name = str(column_meta.get('name') or '')
    label = str(column_meta.get('label') or name)
    comment = str(column_meta.get('comment') or '')
    blob = ' '.join([name.lower(), label.lower(), comment.lower()])
    score = max(_overlap_score_db(query, name), _overlap_score_db(query, label), _overlap_score_db(query, comment))
    reasons: list[str] = []
    if score > 0:
        reasons.append('问题与列名/标签有字面相关')
    hints = _FILTER_NAME_HINTS.get(semantic_type, {})
    strong = tuple(hints.get('strong', tuple()))
    fallback = tuple(hints.get('fallback', tuple()))
    negative = tuple(hints.get('negative', tuple()))
    if strong and any(tok.lower() in blob for tok in strong):
        score += 55
        reasons.append('列名语义与过滤条件强匹配')
    if fallback and any(tok.lower() in blob for tok in fallback):
        score += 16
        reasons.append('列名可能承载该过滤线索')
    if negative and any(tok.lower() in blob for tok in negative):
        score -= 40
        reasons.append('列名更像其它维度，已降权')

    if semantic_type in {'country', 'platform', 'category', 'generic'}:
        if column_meta.get('is_text_like'):
            score += 10
        if column_meta.get('enum_candidate') == 'likely_low':
            score += 4
    if semantic_type == 'time':
        if column_meta.get('is_time_like'):
            score += 60
            reasons.append('时间条件优先落在时间列')
        elif column_meta.get('is_numeric') and any(tok in blob for tok in ('year', 'month', 'day', '年月', '日期', '月')):
            score += 20
            reasons.append('数值列但列名像时间粒度')
        else:
            score -= 15
    if semantic_type == 'id':
        if _is_id_like_name(name):
            score += 60
            reasons.append('列名像 ID/编号键')
        if column_meta.get('primary_key'):
            score += 20
            reasons.append('主键列')
        if column_meta.get('is_numeric'):
            score += 8
        if column_meta.get('is_time_like'):
            score -= 20
    if semantic_type == 'country' and any(tok in blob for tok in ('phone', 'mobile', 'tel', '联系电话', '手机')):
        reasons.append('手机号列可能含国家码，可用于近似口径')
    if semantic_type == 'generic' and column_meta.get('is_time_like'):
        score += 4
    return max(score, 0), reasons


def infer_filter_columns(
    query: str,
    table: str = USERS_TABLE,
    value: str | None = None,
    semantic_type: str | None = None,
    top_k: int = 8,
    preview_limit: int = 6,
) -> dict[str, Any]:
    """综合列名语义 + 小样本值证据，推断某个过滤条件最可能落在哪些列。"""
    table_name = resolve_table_name(table)
    if not is_queryable_table(table_name):
        return {"ok": False, "error": f"Unknown or disallowed table: {table_name}", "table": table_name, "columns": []}
    text = str(query or '').strip()
    if not text:
        return {"ok": False, "error": 'query 不能为空', "table": table_name, "columns": []}
    resolved_semantic = _infer_filter_semantic_type(query=text, value=value, semantic_type=semantic_type)
    rank: list[dict[str, Any]] = []
    for col in list_columns(table_name):
        score, reasons = _column_semantic_match_score(col, f"{text} {value or ''}".strip(), resolved_semantic)
        rank.append(
            {
                'column': col.get('name'),
                'score': score,
                'reasons': reasons,
                'matched_variants': [],
                'match_count': 0,
                'example_values': [],
            }
        )
    rank.sort(key=lambda row: (int(row.get('score') or 0), str(row.get('column') or '')), reverse=True)
    preview_cols = [str(item['column']) for item in rank[: min(max(int(preview_limit), 1), 8)] if item.get('column')]
    variants = _expand_keyword_variants(value=value, query=text, semantic_type=resolved_semantic)
    evidence_map: dict[str, dict[str, Any]] = {}
    if variants and resolved_semantic in {'country', 'platform', 'category', 'generic'} and preview_cols:
        evidence = search_value_examples(
            keyword_variants=variants,
            table=table_name,
            columns=preview_cols,
            semantic_type=resolved_semantic,
            limit_per_column=2,
            max_columns=len(preview_cols),
        )
        for item in evidence.get('columns', []) if isinstance(evidence, dict) else []:
            cname = str(item.get('column') or '')
            if cname:
                evidence_map[cname] = item

    merged: list[dict[str, Any]] = []
    for item in rank:
        cname = str(item.get('column') or '')
        ev = evidence_map.get(cname, {})
        match_count = int(ev.get('total_match_count') or 0)
        score = int(item.get('score') or 0)
        reasons = list(item.get('reasons') or [])
        if match_count > 0:
            score += min(45, 10 + min(match_count, 25))
            reasons.append('列值样本中存在目标值证据')
        merged.append(
            {
                'column': cname,
                'score': score,
                'confidence': 'high' if score >= 90 else ('medium' if score >= 55 else 'low'),
                'reasons': reasons,
                'matched_variants': list(ev.get('matched_variants') or []),
                'match_count': match_count,
                'example_values': list(ev.get('example_values') or []),
            }
        )
    merged.sort(key=lambda row: (int(row.get('score') or 0), int(row.get('match_count') or 0), str(row.get('column') or '')), reverse=True)
    picks = merged[: min(max(int(top_k), 1), 20)]
    return {
        'ok': True,
        'error': None,
        'table': table_name,
        'query': text,
        'value': str(value or '').strip() or None,
        'semantic_type': resolved_semantic,
        'variants': variants,
        'columns': picks,
    }


def infer_count_strategy(
    query: str,
    table: str = USERS_TABLE,
    top_k: int = 5,
) -> dict[str, Any]:
    """为“有多少/计数”类问题推断 COUNT(*) 还是 COUNT(DISTINCT 某键) 更合理。"""
    table_name = resolve_table_name(table)
    if not is_queryable_table(table_name):
        return {"ok": False, "error": f"Unknown or disallowed table: {table_name}", "table": table_name, "candidates": []}
    text = str(query or '').strip()
    if not text:
        return {"ok": False, "error": 'query 不能为空', "table": table_name, "candidates": []}
    row_count = get_table_row_count(table_name)
    q_lower = text.lower()
    entity_hints = {
        'user': ('用户', 'user', 'uid', 'member', 'customer', '客户'),
        'order': ('订单', 'order', '单号', 'plan_no'),
        'material': ('物料', '料号', 'matnr', 'sku', 'item', 'product', '商品'),
    }
    detected_entity = 'record'
    for entity, tokens in entity_hints.items():
        if any(tok.lower() in q_lower for tok in tokens):
            detected_entity = entity
            break

    ranked: list[dict[str, Any]] = []
    for col in list_columns(table_name):
        name = str(col.get('name') or '')
        low = name.lower()
        score = 0
        reasons: list[str] = []
        if col.get('primary_key'):
            score += 40
            reasons.append('主键列')
        if _is_id_like_name(name):
            score += 45
            reasons.append('ID/编号列')
        if col.get('is_numeric'):
            score += 8
        if col.get('is_time_like'):
            score -= 15
        if _is_measure_like_name(name):
            score -= 35
        if detected_entity == 'user' and any(tok in low for tok in ('用户', 'user', 'uid', 'member', 'customer')):
            score += 48
            reasons.append('与“用户”实体匹配')
        if detected_entity == 'order' and any(tok in low for tok in ('order', '订单', 'plan_no', '单号')):
            score += 36
            reasons.append('与“订单”实体匹配')
        if detected_entity == 'material' and any(tok in low for tok in ('matnr', 'sku', 'item', 'product', '物料', '商品')):
            score += 32
            reasons.append('与“物料/商品”实体匹配')
        if detected_entity != 'record' and low == 'id':
            score -= 12
            reasons.append('更像技术主键，已略降权')
        if score > 0:
            ranked.append({'column': name, 'score': score, 'reasons': reasons})
    ranked.sort(key=lambda row: (int(row.get('score') or 0), str(row.get('column') or '')), reverse=True)
    candidates = ranked[: min(max(int(top_k), 1), 8)]

    qtable = _quote_identifier(table_name)
    with get_connection() as conn:
        for item in candidates:
            cname = str(item['column'])
            qcol = _quote_identifier(cname)
            stat = _conn_exec(conn,
                f"""
                SELECT COUNT({qcol}) AS non_null_rows,
                       COUNT(DISTINCT {qcol}) AS distinct_count
                FROM {qtable}
                """
            ).fetchone()
            non_null = int(stat['non_null_rows']) if stat else 0
            distinct_count = int(stat['distinct_count']) if stat else 0
            distinct_ratio = (distinct_count / non_null) if non_null else 0.0
            score = int(item.get('score') or 0)
            reasons = list(item.get('reasons') or [])
            if distinct_ratio >= 0.98:
                score += 25
                reasons.append('去重后几乎唯一')
            elif distinct_ratio >= 0.90:
                score += 12
                reasons.append('高基数键')
            if non_null == row_count and row_count > 0:
                score += 5
            item.update(
                {
                    'non_null_rows': non_null,
                    'distinct_count': distinct_count,
                    'distinct_ratio': round(distinct_ratio, 6),
                    'score': score,
                    'expression': f"COUNT(DISTINCT `{cname}`)",
                    'reasons': reasons,
                }
            )
    candidates.sort(key=lambda row: (int(row.get('score') or 0), float(row.get('distinct_ratio') or 0.0), str(row.get('column') or '')), reverse=True)
    suggested_expression = 'COUNT(*)'
    suggestion_reason = '默认按记录行数统计。'
    if candidates:
        best = candidates[0]
        if detected_entity != 'record' or float(best.get('distinct_ratio') or 0.0) >= 0.9:
            suggested_expression = str(best.get('expression') or suggested_expression)
            suggestion_reason = f"问题更像在统计{detected_entity}数量，`{best.get('column')}` 是当前表最像实体键的列。"
        elif row_count > 0 and int(best.get('distinct_count') or 0) == row_count:
            suggested_expression = 'COUNT(*)'
            suggestion_reason = '当前表近似一行一实体，COUNT(*) 即可。'
    return {
        'ok': True,
        'error': None,
        'table': table_name,
        'query': text,
        'entity_hint': detected_entity,
        'row_count': row_count,
        'row_count_expression': 'COUNT(*)',
        'suggested_expression': suggested_expression,
        'suggestion_reason': suggestion_reason,
        'candidates': candidates,
    }


def validate_join_candidate(
    from_table: str,
    from_column: str,
    to_table: str,
    to_column: str,
    sample_limit: int = 5,
) -> dict[str, Any]:
    """用 distinct overlap 校验某组 JOIN 键是否值得用，避免只按同名列盲 JOIN。"""
    left_table = str(from_table or '').strip()
    right_table = str(to_table or '').strip()
    left_col = str(from_column or '').strip()
    right_col = str(to_column or '').strip()
    if not left_table or not right_table or not left_col or not right_col:
        return {'ok': False, 'error': 'from_table/from_column/to_table/to_column 均不能为空'}
    if not is_queryable_table(left_table) or not is_queryable_table(right_table):
        return {'ok': False, 'error': '表不可查询或不存在'}
    if left_col not in column_names_for_table(left_table) or right_col not in column_names_for_table(right_table):
        return {'ok': False, 'error': '列不存在于对应表中'}

    left_meta = next((c for c in list_columns(left_table) if str(c.get('name') or '') == left_col), {})
    right_meta = next((c for c in list_columns(right_table) if str(c.get('name') or '') == right_col), {})
    qlt = _quote_identifier(left_table)
    qrt = _quote_identifier(right_table)
    qlc = _quote_identifier(left_col)
    qrc = _quote_identifier(right_col)
    safe_limit = min(max(int(sample_limit), 1), 20)
    normalized_time_expr_left = f"SUBSTR(TRIM(CAST({qlc} AS CHAR)), 1, 7)"
    normalized_time_expr_right = f"SUBSTR(TRIM(CAST({qrc} AS CHAR)), 1, 7)"
    looks_time_pair = bool(left_meta.get('is_time_like') and right_meta.get('is_time_like'))
    with get_connection() as conn:
        left_stat = _conn_exec(conn,
            f"SELECT COUNT({qlc}) AS non_null_rows, COUNT(DISTINCT {qlc}) AS distinct_count FROM {qlt}"
        ).fetchone()
        right_stat = _conn_exec(conn,
            f"SELECT COUNT({qrc}) AS non_null_rows, COUNT(DISTINCT {qrc}) AS distinct_count FROM {qrt}"
        ).fetchone()
        overlap_count_row = _conn_exec(
            conn,
            f"""
            SELECT COUNT(*) AS overlap_distinct_keys
            FROM (
                SELECT DISTINCT CAST({qlc} AS CHAR) AS v
                FROM {qlt}
                WHERE {qlc} IS NOT NULL AND TRIM(CAST({qlc} AS CHAR)) != ''
            ) a
            INNER JOIN (
                SELECT DISTINCT CAST({qrc} AS CHAR) AS v
                FROM {qrt}
                WHERE {qrc} IS NOT NULL AND TRIM(CAST({qrc} AS CHAR)) != ''
            ) b ON a.v = b.v
            """,
        ).fetchone()
        sample_rows = _conn_exec(
            conn,
            f"""
            SELECT a.v AS v
            FROM (
                SELECT DISTINCT CAST({qlc} AS CHAR) AS v
                FROM {qlt}
                WHERE {qlc} IS NOT NULL AND TRIM(CAST({qlc} AS CHAR)) != ''
            ) a
            INNER JOIN (
                SELECT DISTINCT CAST({qrc} AS CHAR) AS v
                FROM {qrt}
                WHERE {qrc} IS NOT NULL AND TRIM(CAST({qrc} AS CHAR)) != ''
            ) b ON a.v = b.v
            LIMIT %s
            """,
            (safe_limit,),
        ).fetchall()
        normalized_overlap = 0
        normalized_samples: list[str] = []
        if looks_time_pair:
            normalized_overlap_row = _conn_exec(
                conn,
                f"""
                SELECT COUNT(*) AS overlap_distinct_keys
                FROM (
                    SELECT DISTINCT {normalized_time_expr_left} AS v
                    FROM {qlt}
                    WHERE {qlc} IS NOT NULL AND TRIM(CAST({qlc} AS CHAR)) != ''
                ) a
                INNER JOIN (
                    SELECT DISTINCT {normalized_time_expr_right} AS v
                    FROM {qrt}
                    WHERE {qrc} IS NOT NULL AND TRIM(CAST({qrc} AS CHAR)) != ''
                ) b ON a.v = b.v
                """,
            ).fetchone()
            normalized_overlap = int(normalized_overlap_row['overlap_distinct_keys']) if normalized_overlap_row else 0
            normalized_sample_rows = _conn_exec(
                conn,
                f"""
                SELECT a.v AS v
                FROM (
                    SELECT DISTINCT {normalized_time_expr_left} AS v
                    FROM {qlt}
                    WHERE {qlc} IS NOT NULL AND TRIM(CAST({qlc} AS CHAR)) != ''
                ) a
                INNER JOIN (
                    SELECT DISTINCT {normalized_time_expr_right} AS v
                    FROM {qrt}
                    WHERE {qrc} IS NOT NULL AND TRIM(CAST({qrc} AS CHAR)) != ''
                ) b ON a.v = b.v
                LIMIT %s
                """,
                (safe_limit,),
            ).fetchall()
            normalized_samples = [row['v'] for row in normalized_sample_rows]
    left_distinct = int(left_stat['distinct_count']) if left_stat else 0
    right_distinct = int(right_stat['distinct_count']) if right_stat else 0
    overlap = int(overlap_count_row['overlap_distinct_keys']) if overlap_count_row else 0
    effective_overlap = max(overlap, normalized_overlap)
    coverage_left = (effective_overlap / left_distinct) if left_distinct else 0.0
    coverage_right = (effective_overlap / right_distinct) if right_distinct else 0.0
    confidence = 'low'
    if effective_overlap > 0 and (coverage_left >= 0.2 or coverage_right >= 0.2):
        confidence = 'high'
    elif effective_overlap > 0 and (coverage_left >= 0.05 or coverage_right >= 0.05):
        confidence = 'medium'
    join_transform_hint = None
    if overlap == 0 and normalized_overlap > 0 and looks_time_pair:
        join_transform_hint = f"原值格式不一致，可按年月对齐：SUBSTR(`{left_table}`.`{left_col}`,1,7) = SUBSTR(`{right_table}`.`{right_col}`,1,7)"
    return {
        'ok': True,
        'error': None,
        'from_table': left_table,
        'from_column': left_col,
        'to_table': right_table,
        'to_column': right_col,
        'on': f"`{left_table}`.`{left_col}` = `{right_table}`.`{right_col}`",
        'left_stats': {
            'non_null_rows': int(left_stat['non_null_rows']) if left_stat else 0,
            'distinct_count': left_distinct,
        },
        'right_stats': {
            'non_null_rows': int(right_stat['non_null_rows']) if right_stat else 0,
            'distinct_count': right_distinct,
        },
        'overlap_distinct_keys': overlap,
        'normalized_overlap_distinct_keys': normalized_overlap if looks_time_pair else None,
        'coverage_left': round(coverage_left, 6),
        'coverage_right': round(coverage_right, 6),
        'confidence': confidence,
        'sample_overlap_keys': [row['v'] for row in sample_rows],
        'normalized_sample_overlap_keys': normalized_samples if looks_time_pair else [],
        'join_transform_hint': join_transform_hint,
    }


def _keyword_probe_target_columns(
    columns_meta: list[dict[str, Any]],
    keyword: str,
    column_names: list[str] | None,
) -> tuple[list[str], dict[str, Any]]:
    """
    未指定 column_names 时：非纯数字关键词只扫 is_text_like（国家/地名不会在纯数值列）；
    长数字串（如电话/ID）再扫 is_text_like + is_numeric。再按 USER_RAG_KEYWORD_MAX_COLUMNS 截断。
    """
    names_in_order = [c["name"] for c in columns_meta]
    name_set = set(names_in_order)
    max_c = USER_RAG_KEYWORD_MAX_COLUMNS
    meta: dict[str, Any] = {
        "max_columns_per_call": max_c,
        "truncated": False,
        "scope": "explicit",
    }
    if column_names:
        targets = [c for c in column_names if c in name_set]
        if not targets:
            targets = list(names_in_order)
            meta["scope"] = "fallback_all"
        else:
            meta["scope"] = "explicit"
    else:
        digitish = bool(re.fullmatch(r"\d{3,}", str(keyword or "").strip()))
        if digitish:
            targets = [
                c["name"]
                for c in columns_meta
                if c.get("is_text_like") or c.get("is_numeric")
            ]
            meta["scope"] = "digit_text_or_numeric"
        else:
            targets = [c["name"] for c in columns_meta if c.get("is_text_like")]
            meta["scope"] = "text_like_only"
        if not targets:
            targets = list(names_in_order)
            meta["scope"] = "fallback_all"
    total_before_cap = len(targets)
    if len(targets) > max_c:
        targets = targets[:max_c]
        meta["truncated"] = True
    meta["candidate_columns"] = total_before_cap
    meta["scanned_columns"] = len(targets)
    return targets, meta


def search_keyword_across_columns(
    keyword: str,
    column_names: list[str] | None = None,
    table: str = USERS_TABLE,
) -> dict[str, Any]:
    """
    在多个文本列中统计包含 keyword 的行数，并各取一条样例值。
    用于发现「国家名、地名」等是否藏在 `用户昵称` 等列，而不是仅看 `目标用户标签`。
    大表默认只扫文本类列且限制列数，避免 60+ 列 × 全表 COUNT 拖死请求。
    """
    text = str(keyword or "").strip()
    table_name = str(table or "").strip() or USERS_TABLE
    if not text:
        return {"error": "keyword 不能为空"}
    if not is_queryable_table(table_name):
        return {"table": table_name, "keyword": text, "matches": []}
    columns_meta = list_columns(table_name)
    if not columns_meta:
        return {"table": table_name, "keyword": text, "matches": []}
    targets, probe_meta = _keyword_probe_target_columns(columns_meta, text, column_names)
    needle = text.lower()
    matches: list[dict[str, Any]] = []
    qtable = _quote_identifier(table_name)
    with get_connection() as conn:
        for col in targets:
            quoted = _quote_identifier(col)
            count_row = _conn_exec(conn,
                f"""
                SELECT COUNT(*) AS cnt
                FROM {qtable}
                WHERE {quoted} IS NOT NULL
                  AND TRIM(CAST({quoted} AS CHAR)) != ''
                  AND INSTR(LOWER(CAST({quoted} AS CHAR)), %s) > 0
                """,
                (needle,),
            ).fetchone()
            cnt = int(count_row["cnt"]) if count_row else 0
            example: str | None = None
            if cnt > 0:
                ex = _conn_exec(conn,
                    f"""
                    SELECT CAST({quoted} AS CHAR) AS v
                    FROM {qtable}
                    WHERE {quoted} IS NOT NULL
                      AND INSTR(LOWER(CAST({quoted} AS CHAR)), %s) > 0
                    LIMIT 1
                    """,
                    (needle,),
                ).fetchone()
                example = ex["v"] if ex else None
            matches.append({"column": col, "match_count": cnt, "example_value": example})
    matches.sort(key=lambda item: item["match_count"], reverse=True)
    return {"table": table_name, "keyword": text, "matches": matches, "meta": probe_meta}


def search_keyword_in_tables(
    keyword: str,
    tables: list[str] | None = None,
    limit_per_table: int = 8,
    positive_only: bool = True,
) -> dict[str, Any]:
    """跨多张表搜 keyword，快速定位“值在哪张表、哪几列”。"""
    text = str(keyword or "").strip()
    if not text:
        return {"ok": False, "error": "keyword 不能为空", "keyword": text, "tables": []}
    allowed = set(queryable_table_names())
    selected: list[str] = []
    raw_tables = tables or queryable_table_names()
    for raw in raw_tables:
        name = str(raw or "").strip()
        if name and name in allowed and name not in selected:
            selected.append(name)
    if not selected:
        return {"ok": False, "error": "没有可查询的目标表", "keyword": text, "tables": []}

    limit_per_table = min(max(int(limit_per_table), 1), 20)
    out: list[dict[str, Any]] = []
    for table_name in selected:
        res = search_keyword_across_columns(keyword=text, table=table_name)
        matches = list(res.get("matches", []) or [])
        positive = [m for m in matches if int(m.get("match_count") or 0) > 0]
        total_match_count = sum(int(m.get("match_count") or 0) for m in positive)
        if positive_only and total_match_count == 0:
            continue
        out.append(
            {
                "table": table_name,
                "total_match_count": total_match_count,
                "matched_columns": positive[:limit_per_table] if positive else matches[:limit_per_table],
                "meta": res.get("meta", {}),
            }
        )
    out.sort(key=lambda item: (item.get("total_match_count", 0), item.get("table", "")), reverse=True)
    return {
        "ok": True,
        "error": None,
        "keyword": text,
        "tables": out,
        "meta": {
            "searched_tables": selected,
            "returned_tables": len(out),
            "positive_only": bool(positive_only),
        },
    }


def profile_table_columns(sample_per_column: int = 3, table: str = USERS_TABLE) -> dict[str, Any]:
    """多列总览（原 columns_data_overview），统一带 ok 信封。"""
    table_name = resolve_table_name(table)
    if not is_queryable_table(table_name):
        return {
            "ok": False,
            "error": f"Unknown or disallowed table: {table_name}",
            "table": table_name,
            "data": None,
        }
    safe_samples = min(max(sample_per_column, 1), 10)
    columns_meta = list_columns(table_name)
    max_cols = USER_RAG_PROFILE_TABLE_MAX_COLUMNS
    total_cols = len(columns_meta)
    profile_truncated = total_cols > max_cols
    if profile_truncated:
        columns_meta = columns_meta[:max_cols]
    overview: list[dict[str, Any]] = []
    qtable = _quote_identifier(table_name)
    with get_connection() as conn:
        for col in columns_meta:
            name = col["name"]
            quoted = _quote_identifier(name)
            stat = _conn_exec(conn,
                f"""
                SELECT COUNT(*) AS total_rows,
                       COUNT({quoted}) AS non_null_rows,
                       COUNT(DISTINCT {quoted}) AS distinct_count
                FROM {qtable}
                """
            ).fetchone()
            samples = _conn_exec(
                conn,
                f"""
                SELECT CAST({quoted} AS CHAR) AS v
                FROM {qtable}
                WHERE {quoted} IS NOT NULL AND TRIM(CAST({quoted} AS CHAR)) != ''
                GROUP BY CAST({quoted} AS CHAR)
                ORDER BY COUNT(*) DESC
                LIMIT %s
                """,
                (safe_samples,),
            ).fetchall()
            overview.append(
                {
                    "column": name,
                    "sql_type": col.get("sql_type", "TEXT"),
                    "non_null_rows": int(stat["non_null_rows"]),
                    "distinct_count": int(stat["distinct_count"]),
                    "sample_values": [row["v"] for row in samples],
                }
            )
    return {
        "ok": True,
        "error": None,
        "table": table_name,
        "data": {
            "columns": overview,
            "meta": {
                "profiled_columns": len(overview),
                "total_table_columns": total_cols,
                "truncated": profile_truncated,
                "max_columns": max_cols,
            },
        },
    }


def columns_data_overview(sample_per_column: int = 3, table: str = USERS_TABLE) -> dict[str, Any]:
    """兼容旧名：同 profile_table_columns，返回扁平 columns 列表。"""
    r = profile_table_columns(sample_per_column=sample_per_column, table=table)
    if not r.get("ok"):
        return {"table": r.get("table", table), "columns": []}
    cols = (r.get("data") or {}).get("columns", [])
    return {"table": r["table"], "columns": cols}


def sample_rows(table: str, limit: int = 5) -> dict[str, Any]:
    """兼容：等价于 inspect_rows(as_dict=True)。"""
    r = inspect_rows(table=table, limit=limit, as_dict=True)
    if not r.get("ok"):
        return {"error": r.get("error"), "table": r.get("table"), "rows": []}
    return {"table": r["table"], "rows": r["rows"]}


def find_relevant_columns(query: str, table: str = USERS_TABLE, top_k: int = 5) -> dict[str, Any]:
    """按问题与列名/注释字面相关度排序，辅助选列（非值域搜索）。"""
    table_name = resolve_table_name(table)
    if not is_queryable_table(table_name):
        return {"ok": False, "error": f"Unknown or disallowed table: {table_name}", "table": table_name, "data": None}
    q = str(query or "").strip()
    top_k = min(max(int(top_k), 1), 20)
    ranked: list[tuple[int, str, list[str]]] = []
    for col in list_columns(table_name):
        name = col["name"]
        label = str(col.get("label") or name)
        comment = str(col.get("comment") or "")
        sn = _overlap_score_db(q, name)
        sl = _overlap_score_db(q, label)
        sc = _overlap_score_db(q, comment)
        score = max(sn, sl, sc)
        reasons: list[str] = []
        if sn > 0:
            reasons.append("列名命中")
        if sl > 0 and label != name:
            reasons.append("标签命中")
        if sc > 0 and comment:
            reasons.append("注释命中")
        if col.get("is_time_like") and any(k in q for k in ("时间", "日期", "天", "月", "年", "最近", "活跃")):
            score += 2
            reasons.append("时间语义列+问题含时间词")
        ranked.append((score, name, reasons))
    ranked.sort(key=lambda x: x[0], reverse=True)
    picks = [{"column": n, "score": s, "reasons": rs} for s, n, rs in ranked if s > 0][:top_k]
    if not picks:
        picks = [{"column": n, "score": s, "reasons": rs or ["默认前列"]} for s, n, rs in ranked[:top_k]]
    return {"ok": True, "error": None, "table": table_name, "data": {"columns": picks}}


def find_time_columns(table: str = USERS_TABLE) -> dict[str, Any]:
    """根据类型与列名启发式列出疑似时间列。"""
    table_name = resolve_table_name(table)
    if not is_queryable_table(table_name):
        return {"ok": False, "error": f"Unknown or disallowed table: {table_name}", "table": table_name, "data": None}
    out: list[dict[str, Any]] = []
    for col in list_columns(table_name):
        if not col.get("is_time_like"):
            continue
        nm = col["name"].lower()
        strong = any(k in nm for k in ("时间", "日期", "time", "date", "created", "updated"))
        out.append(
            {
                "column": col["name"],
                "confidence": "high" if strong else "medium",
                "reasons": ["sql_type 或列名含时间语义"],
            }
        )
    return {"ok": True, "error": None, "table": table_name, "data": {"time_columns": out}}


def find_join_path(from_table: str, to_table: str) -> dict[str, Any]:
    """优先走外键 BFS；无 FK 时回退到启发式同名键/时间键 JOIN 候选。"""
    ft = str(from_table or "").strip()
    tt = str(to_table or "").strip()
    if not ft or not tt:
        return {"ok": False, "error": "from_table 与 to_table 均不能为空", "data": None}
    if not is_queryable_table(ft) or not is_queryable_table(tt):
        return {"ok": False, "error": "表不可查询或不存在", "data": None}
    if ft == tt:
        return {"ok": True, "error": None, "data": {"path": [ft], "joins": [], "note": "同表无需 JOIN"}}

    def _bfs(adj: dict[str, list[tuple[str, dict[str, str]]]]) -> tuple[list[str], list[dict[str, str]]] | None:
        q = deque([ft])
        prev: dict[str, str | None] = {ft: None}
        how: dict[str, dict[str, str]] = {}
        while q:
            u = q.popleft()
            if u == tt:
                break
            for v, meta in adj.get(u, []):
                if v not in prev:
                    prev[v] = u
                    how[v] = meta
                    q.append(v)
        if tt not in prev:
            return None
        nodes: list[str] = []
        cur: str | None = tt
        while cur is not None:
            nodes.append(cur)
            cur = prev[cur]
        nodes.reverse()
        joins = [how[nodes[i]] for i in range(1, len(nodes))]
        return nodes, joins

    rels = get_table_relationships()
    adj_fk: dict[str, list[tuple[str, dict[str, str]]]] = {}
    for r in rels:
        a, b = r["from_table"], r["to_table"]
        fc, tc = r["from_column"], r["to_column"]
        meta = {
            "on": f'{_quote_identifier(a)}.{_quote_identifier(fc)} = {_quote_identifier(b)}.{_quote_identifier(tc)}',
            "via": "foreign_key",
        }
        adj_fk.setdefault(a, []).append((b, meta))
        adj_fk.setdefault(b, []).append((a, meta))
    fk_path = _bfs(adj_fk)
    if fk_path is not None:
        nodes, joins = fk_path
        return {"ok": True, "error": None, "data": {"path": nodes, "joins": joins, "note": "基于外键路径"}}

    heur = infer_join_candidates(tables=queryable_table_names(), limit=200)
    adj_h: dict[str, list[tuple[str, dict[str, str]]]] = {}
    for item in heur.get("candidates", []):
        a, b = item["from_table"], item["to_table"]
        meta = {
            "on": item["on"],
            "via": "heuristic_shared_key",
            "confidence": item.get("confidence", "medium"),
        }
        adj_h.setdefault(a, []).append((b, meta))
        adj_h.setdefault(b, []).append((a, meta))
    heuristic_path = _bfs(adj_h)
    if heuristic_path is not None:
        nodes, joins = heuristic_path
        return {
            "ok": True,
            "error": None,
            "data": {"path": nodes, "joins": joins, "note": "未声明外键，已回退到启发式同名键/时间键路径"},
        }

    return {
        "ok": True,
        "error": None,
        "data": {"path": [], "joins": [], "note": "未找到外键或高置信启发式 JOIN 路径"},
    }


def search_similar_values(
    field: str,
    query: str,
    limit: int = 10,
    table: str = USERS_TABLE,
) -> dict[str, Any]:
    """对字段已有取值做模糊/编辑距离排序，纠错用户输入（如拼写）。"""
    table_name = resolve_table_name(table)
    q = str(query or "").strip()
    if not q:
        return {"ok": False, "error": "query 不能为空", "table": table_name, "data": None}
    names = set(column_names_for_table(table_name))
    if field not in names:
        return {"ok": False, "error": f"Unknown field: {field}", "table": table_name, "data": None}
    quoted = _quote_identifier(field)
    qtable = _quote_identifier(table_name)
    with get_connection() as conn:
        rows = _conn_exec(conn,
            f"""
            SELECT {quoted} AS value, COUNT(*) AS cnt
            FROM {qtable}
            WHERE {quoted} IS NOT NULL AND TRIM(CAST({quoted} AS CHAR)) != ''
            GROUP BY {quoted}
            ORDER BY cnt DESC
            LIMIT 200
            """
        ).fetchall()
    qlow = q.lower()
    scored: list[tuple[float, Any, int]] = []
    for row in rows:
        val = row["value"]
        vs = str(val).strip()
        if not vs:
            continue
        vl = vs.lower()
        ratio = difflib.SequenceMatcher(None, qlow, vl).ratio()
        bonus = 3.0 if qlow in vl or vl in qlow else 0.0
        scored.append((ratio * 10 + bonus + min(int(row["cnt"]) / 5000.0, 0.3), val, int(row["cnt"])))
    scored.sort(key=lambda x: x[0], reverse=True)
    lim = min(max(int(limit), 1), 30)
    matches = [{"value": v, "count": c, "score": round(s, 4)} for s, v, c in scored[:lim]]
    return {"ok": True, "error": None, "table": table_name, "data": {"field": field, "matches": matches}}


def summarize_query_result(sql: str, max_rows: int = 500) -> dict[str, Any]:
    """对（截断后的）查询结果做轻量统计，便于 agent 自检。"""
    cap = min(max(int(max_rows), 1), 2000)
    try:
        res = execute_sql(sql, max_rows=cap)
    except ValueError as exc:
        return {"ok": False, "error": str(exc), "data": None}
    cols: list[str] = res["columns"]
    rows: list[list[Any]] = res["rows"]
    n = len(rows)
    per_col: dict[str, Any] = {}
    for ci, cname in enumerate(cols):
        col_vals = [r[ci] if ci < len(r) else None for r in rows]
        non_null = [v for v in col_vals if v is not None and str(v).strip() != ""]
        if not non_null:
            per_col[cname] = {"kind": "empty", "non_null": 0}
            continue
        nums: list[float] = []
        for v in non_null:
            if isinstance(v, bool):
                continue
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                nums.append(float(v))
            else:
                try:
                    nums.append(float(str(v)))
                except ValueError:
                    break
        if len(nums) == len(non_null):
            per_col[cname] = {
                "kind": "numeric",
                "min": min(nums),
                "max": max(nums),
                "avg": sum(nums) / len(nums),
                "non_null": len(non_null),
            }
        else:
            freq: dict[str, int] = {}
            for v in non_null:
                k = str(v)[:200]
                freq[k] = freq.get(k, 0) + 1
            top = sorted(freq.items(), key=lambda x: x[1], reverse=True)[:5]
            per_col[cname] = {"kind": "categorical", "distinct_in_sample": len(freq), "top_values": top}
    return {
        "ok": True,
        "error": None,
        "data": {
            "row_count": n,
            "truncated": bool(res.get("truncated")),
            "columns": cols,
            "column_summaries": per_col,
        },
    }


def _strip_trailing_sql_noise(sql: str) -> str:
    """去掉句末句号，避免 ...'2026-02-03'. LIMIT 触发 near LIMIT 语法错误。"""
    s = sql.strip().rstrip(";")
    while s and s[-1] in (".", "。", "．", "·"):
        s = s[:-1].rstrip()
    return s.strip()


def _guard_select_sql(sql: str) -> str:
    normalized = " ".join(_strip_trailing_sql_noise(sql).split())
    lowered = normalized.lower()
    if not lowered.startswith("select "):
        raise ValueError("Only SELECT SQL is allowed.")
    if ";" in normalized:
        raise ValueError("Only one SQL statement is allowed.")
    if any(re.search(rf"\b{re.escape(keyword)}\b", lowered) for keyword in FORBIDDEN_SQL):
        raise ValueError("Unsafe SQL detected.")
    tables = {
        token
        for pair in re.findall(r"\bfrom\s+([a-zA-Z_][a-zA-Z0-9_]*)|\bjoin\s+([a-zA-Z_][a-zA-Z0-9_]*)", lowered)
        for token in pair
        if token
    }
    allowed = set(queryable_table_names())
    if tables and not tables.issubset(allowed):
        bad = sorted(tables - allowed)
        raise ValueError(
            "SQL 只能查询当前库内的业务表（不含日志等系统表）。"
            + (f" 不允许的表: {', '.join(bad)}" if bad else "")
        )
    return normalized


def validate_sql(sql: str) -> dict[str, Any]:
    """只读校验：不执行查询；通过时返回规范化后的 SQL 文本与风险等级。"""
    try:
        normalized = _guard_select_sql(sql)
    except ValueError as exc:
        return {
            "ok": False,
            "error": str(exc),
            "errors": [str(exc)],
            "warnings": [],
            "normalized_sql": None,
            "risk_level": "high",
        }
    warnings: list[str] = []
    lowered = normalized.lower()
    if re.search(r"\bselect\s+\*", lowered):
        warnings.append("使用了 SELECT *，高基数传输与语义不明确风险较高。")
    has_group = bool(re.search(r"\bgroup\s+by\b", lowered))
    has_agg = bool(re.search(r"\b(count|sum|avg|min|max)\s*\(", lowered, re.IGNORECASE))
    if not re.search(r"\bwhere\b", lowered) and not has_group and not has_agg:
        warnings.append("未检测到 WHERE；可能是全表明细扫描。")
    if not re.search(r"\blimit\s+\d+", lowered):
        warnings.append("未检测到 LIMIT；大范围结果集时注意性能与内存。")
    if not re.search(r"\border\s+by\b", lowered) and not re.search(r"\blimit\s+\d+", lowered) and not has_agg:
        warnings.append("无 ORDER BY 且无 LIMIT；返回顺序不稳定。")
    unknown_cols = _validate_sql_column_identifiers(normalized, lowered)
    for u in unknown_cols[:12]:
        warnings.append(f"标识符 `{u}` 不在已加载表列中（若用了别名可忽略）。")

    risk_level = "low"
    if any("SELECT *" in w or "高基数" in w or "全表" in w for w in warnings):
        risk_level = "high"
    elif warnings:
        risk_level = "medium"

    return {
        "ok": True,
        "error": None,
        "errors": [],
        "warnings": warnings,
        "normalized_sql": normalized,
        "risk_level": risk_level,
    }


# WHERE 语义：国家/地名类子串不应出现在时间列、渠道列、纯 ID 列的条件里（易致 0 行或逻辑错）
_BUILTIN_COUNTRY_SQL_FRAGMENTS: frozenset[str] = frozenset(
    {
        "indonesia",
        "indonesian",
        "india",
        "indian",
        "thailand",
        "vietnam",
        "malaysia",
        "philippines",
        "china",
        "chinese",
        "japan",
        "korea",
        "usa",
        "america",
        "brazil",
        "印度尼西亚",
        "印尼",
        "印度",
        "泰国",
        "越南",
        "马来西亚",
        "菲律宾",
        "中国",
        "日本",
        "韩国",
        "美国",
        "巴西",
    }
)


def _country_tokens_for_semantics_check(question: str | None) -> set[str]:
    out: set[str] = set(_BUILTIN_COUNTRY_SQL_FRAGMENTS)
    q = (question or "").strip()
    if not q:
        return out
    intent = resolve_country_intent(q)
    if intent:
        out.add(intent.canonical_en.strip().lower())
        out.add(intent.cn_display.strip().lower())
        for tr in intent.triggers:
            if tr.strip():
                out.add(tr.strip().lower())
    return {t for t in out if t}


def _where_clause_body(sql: str) -> str:
    """取出顶层 WHERE 与 ORDER/GROUP/LIMIT 之间的片段（忽略字符串内的括号）。"""
    m = re.search(r"\bWHERE\b", sql, re.IGNORECASE)
    if not m:
        return ""
    start = m.end()
    i = start
    depth = 0
    n = len(sql)
    while i < n:
        c = sql[i]
        if c == "'":
            i += 1
            while i < n:
                if sql[i] == "'":
                    if i + 1 < n and sql[i + 1] == "'":
                        i += 2
                        continue
                    i += 1
                    break
                i += 1
            continue
        if c == "(":
            depth += 1
        elif c == ")":
            depth = max(0, depth - 1)
        elif depth == 0 and re.match(r"\s+(?:ORDER|GROUP)\s+BY\b|\s+LIMIT\b", sql[i:], re.IGNORECASE):
            break
        i += 1
    return sql[start:i].strip()


def _sql_chunk_after_identifier(sql: str, needle: str, max_len: int = 480) -> str:
    i = sql.find(needle)
    if i < 0:
        return ""
    return sql[i : i + len(needle) + max_len]


def _chunk_has_string_predicate_with_tokens(chunk: str, tokens: set[str]) -> bool:
    """片段内出现 LIKE/INSTR/GLOB 且字面量侧含某国别相关子串（大小写不敏感 + 中文原样）。"""
    cl = chunk.lower()
    if "like" not in cl and "instr" not in cl and "glob" not in cl:
        return False
    for tok in tokens:
        if not tok:
            continue
        tl = tok.lower()
        if tl.isascii():
            if tl in cl:
                return True
        elif tok in chunk:
            return True
    return False


def _is_platform_channel_column(col: dict[str, Any]) -> bool:
    name = str(col.get("name") or "")
    return name == "所属平台" or ("平台" in name and "渠道" not in name and len(name) <= 12)


def _is_likely_id_column(col: dict[str, Any]) -> bool:
    name = str(col.get("name") or "")
    nl = name.lower()
    if col.get("is_numeric") and ("id" in nl or "编号" in name or name.endswith("ID") or name == "用户ID"):
        return True
    return name in ("用户ID", "id", "ID")


def validate_where_predicate_column_fit(sql: str, question: str | None = None) -> dict[str, Any]:
    """
    检查 SELECT 的 WHERE 中是否把「国家/地名」类筛选错绑到时间、渠道、ID 等列上。
    不解析完整 SQL AST，用列元数据 + 局部子串启发式；供 Agent 工具与 Finish 守卫使用。
    """
    try:
        normalized = _guard_select_sql(sql)
    except ValueError as exc:
        return {
            "ok": False,
            "predicate_fit_ok": False,
            "error": str(exc),
            "issues": [],
            "normalized_sql": None,
            "hint": "先修正 SQL 语法与安全限制后再做语义检查。",
        }
    tokens = _country_tokens_for_semantics_check(question)
    lowered = normalized.lower()
    tables: set[str] = set()
    for pair in re.findall(r"\bfrom\s+([a-zA-Z_][a-zA-Z0-9_]*)|\bjoin\s+([a-zA-Z_][a-zA-Z0-9_]*)", lowered):
        for token in pair:
            if token:
                tables.add(token)
    for pair in re.findall(r"\bfrom\s+`([a-zA-Z_][a-zA-Z0-9_]*)`|\bjoin\s+`([a-zA-Z_][a-zA-Z0-9_]*)`", lowered):
        for token in pair:
            if token:
                tables.add(token)
    allowed_names = {t.lower(): t for t in queryable_table_names()}
    resolved_tables = [allowed_names[t] for t in tables if t in allowed_names]
    if not resolved_tables:
        return {
            "ok": True,
            "predicate_fit_ok": True,
            "error": None,
            "issues": [],
            "normalized_sql": normalized,
            "hint": "未解析到 FROM 表，跳过列语义绑定检查。",
        }
    where_body = _where_clause_body(normalized)
    if not where_body.strip():
        return {
            "ok": True,
            "predicate_fit_ok": True,
            "error": None,
            "issues": [],
            "normalized_sql": normalized,
            "hint": "无 WHERE 子句，跳过列与筛选条件绑定检查。",
        }
    issues: list[dict[str, Any]] = []
    for table in resolved_tables:
        cols = list_columns(table)
        for col in cols:
            cname = str(col.get("name") or "")
            if not cname:
                continue
            needles = [f"`{table}`.`{cname}`", f"`{cname}`"]
            is_time = bool(col.get("is_time_like"))
            if not is_time and re.search(r"(时间|日期|timestamp|date|time)$", cname, re.IGNORECASE):
                is_time = True
            for needle in needles:
                if needle not in where_body:
                    continue
                chunk = _sql_chunk_after_identifier(where_body, needle)
                if not _chunk_has_string_predicate_with_tokens(chunk, tokens):
                    continue
                if is_time:
                    issues.append(
                        {
                            "severity": "high",
                            "code": "country_like_on_time_column",
                            "table": table,
                            "column": cname,
                            "detail": f"列 `{table}`.`{cname}` 为时间/日期语义，WHERE 中却对国家/地名使用 LIKE/INSTR，"
                            "易错且常得 0 行；国家条件应只放在 Search_keyword 已验证的**昵称/标签等文本列**（见 Knowledge）。",
                        }
                    )
                elif _is_platform_channel_column(col):
                    issues.append(
                        {
                            "severity": "high",
                            "code": "country_like_on_platform_column",
                            "table": table,
                            "column": cname,
                            "detail": f"`{cname}` 表示 APP/PC 等渠道，不要对国家名做 LIKE；请删除该条件或改用正确列。",
                        }
                    )
                elif _is_likely_id_column(col):
                    issues.append(
                        {
                            "severity": "high",
                            "code": "country_like_on_id_column",
                            "table": table,
                            "column": cname,
                            "detail": f"列 `{cname}` 为数值/ID 语义，不宜用国家名字符串 LIKE；请删除或改到合适的文本列。",
                        }
                    )
                break
    high = [x for x in issues if x.get("severity") == "high"]
    return {
        "ok": True,
        "predicate_fit_ok": len(high) == 0,
        "error": None,
        "issues": issues,
        "normalized_sql": normalized,
        "hint": "若有 high 级别 issue，Finish 前应删错列条件或改用 Search_keyword 命中的列；可用本工具自检。",
    }


def explain_sql(sql: str) -> dict[str, Any]:
    """MySQL EXPLAIN，不返回业务数据。"""
    check = validate_sql(sql)
    if not check["ok"]:
        return {**check, "plan": []}
    normalized = check["normalized_sql"]
    if not normalized:
        return check
    with get_connection() as conn:
        plan_rows = _conn_exec(conn, f"EXPLAIN {normalized}").fetchall()
    return {
        "ok": True,
        "error": None,
        "errors": [],
        "warnings": check.get("warnings", []),
        "normalized_sql": normalized,
        "risk_level": check.get("risk_level", "low"),
        "plan": [dict(row) for row in plan_rows],
    }


def execute_sql(sql_query: str, max_rows: int | None = None) -> dict[str, Any]:
    """
    max_rows:
        None — 使用环境变量默认上限（防 OOM）
        <0 — 不截断（仅用于可信导出路径）
    """
    safe_sql = _guard_select_sql(sql_query)
    unlimited = max_rows is not None and max_rows < 0
    cap = EXECUTE_SQL_MAX_ROWS if max_rows is None else max_rows
    if unlimited:
        with get_connection() as conn:
            rows = _conn_exec(conn, safe_sql).fetchall()
        truncated = False
    else:
        cap = min(max(int(cap), 1), 500_000)
        with get_connection() as conn:
            cur = _conn_exec(conn, safe_sql)
            fetched = cur.fetchmany(cap + 1)
        truncated = len(fetched) > cap
        rows = fetched[:cap]
    if not rows:
        return {
            "ok": True,
            "sql": safe_sql,
            "columns": [],
            "rows": [],
            "row_count": 0,
            "truncated": False,
        }
    columns = list(rows[0].keys())
    return {
        "ok": True,
        "sql": safe_sql,
        "columns": columns,
        "rows": [_row_to_ordered_values(row, columns) for row in rows],
        "row_count": len(rows),
        "truncated": truncated,
    }


def export_sql_to_excel(sql_query: str, output_name: str = "result.xlsx") -> dict[str, Any]:
    result = execute_sql(sql_query, max_rows=-1)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "result"
    if result["columns"]:
        sheet.append(result["columns"])
    for row in result["rows"]:
        sheet.append(row)
    file_path = DATA_DIR / output_name
    workbook.save(file_path)
    return {"file_path": str(file_path), "row_count": len(result["rows"]), "sql": result["sql"]}


def export_sql_as_bytes(sql_query: str, fmt: str = "csv") -> tuple[bytes, str, str]:
    """Re-run guarded SELECT and return (body, media_type, filename). fmt: csv | xlsx."""
    key = (fmt or "csv").lower().strip()
    if key not in ("csv", "xlsx"):
        raise ValueError("format 必须是 csv 或 xlsx")
    result = execute_sql(sql_query, max_rows=-1)
    columns: list[str] = result["columns"]
    rows: list[list[Any]] = result["rows"]
    if key == "csv":
        buffer = StringIO()
        writer = csv.writer(buffer)
        if columns:
            writer.writerow(columns)
        for row in rows:
            writer.writerow(row)
        data = buffer.getvalue().encode("utf-8-sig")
        return data, "text/csv; charset=utf-8", "query_result.csv"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "result"
    if columns:
        sheet.append(columns)
    for row in rows:
        sheet.append(row)
    bio = BytesIO()
    workbook.save(bio)
    return bio.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "query_result.xlsx"


def recent_query_examples(limit: int = 5) -> list[dict[str, Any]]:
    """仅用于 RAG：只取用户明确标记为「有帮助」(user_feedback=1) 且含 SQL 的记录。"""
    safe_limit = min(max(limit, 1), 20)
    with get_connection() as conn:
        rows = _conn_exec(
            conn,
            """
            SELECT `question`, `understanding`, `sql`, `created_at`
            FROM `query_logs`
            WHERE `sql` IS NOT NULL AND TRIM(`sql`) <> ''
              AND `user_feedback` = 1
            ORDER BY `id` DESC
            LIMIT %s
            """,
            (safe_limit,),
        ).fetchall()
    return [dict(row) for row in rows]


def log_query(question: str, understanding: str, sql: str | None, tool_trace: list[dict[str, Any]]) -> int:
    """写入一条查询记录，user_feedback 默认为 NULL（未评价），不进入 few-shot 示例。"""
    with get_connection() as conn:
        cursor = _conn_exec(
            conn,
            """
            INSERT INTO `query_logs` (`question`, `understanding`, `sql`, `tool_trace`, `created_at`, `user_feedback`)
            VALUES (%s, %s, %s, %s, %s, NULL)
            """,
            (question, understanding, sql, json.dumps(tool_trace, ensure_ascii=False), datetime.utcnow().isoformat()),
        )
        conn.commit()
        return int(cursor.lastrowid)


import_xlsx_to_sqlite = import_tabular_to_mysql  # 兼容旧导入名


def set_query_feedback(log_id: int, helpful: bool) -> dict[str, Any]:
    """用户反馈：helpful=True -> 1（纳入示例），False -> -1（不纳入）。"""
    if log_id < 1:
        return {"error": "无效的 log_id"}
    value = 1 if helpful else -1
    with get_connection() as conn:
        row = _conn_exec(conn, "SELECT `id` FROM `query_logs` WHERE `id` = %s", (log_id,)).fetchone()
        if not row:
            return {"error": "记录不存在"}
        _conn_exec(conn, "UPDATE `query_logs` SET `user_feedback` = %s WHERE `id` = %s", (value, log_id))
        conn.commit()
    return {"ok": True, "id": log_id, "user_feedback": value}
